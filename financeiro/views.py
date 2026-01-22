from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q, F, Case, When, DecimalField, Count
from django.utils import timezone
from datetime import timedelta, date, datetime
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from decimal import Decimal
import json

from usuarios.decorators import permissao_menu_required
from cadastros.models import PlanoConta, ContaFinanceira, MetodoPagamento
from .models import ContaPagar, ContaReceber, MovimentacaoFinanceira
from .forms import ContaPagarForm, FiltroContaPagarForm, DarBaixaForm, FiltroRelatorioForm


@login_required
@permissao_menu_required('financeiro', 'view')
def dashboard_financeiro(request):
    """Dashboard do módulo financeiro com estatísticas"""
    
    # Estatísticas gerais - Contas a Pagar
    total_pagar = ContaPagar.objects.aggregate(
        total=Sum('valor')
    )['total'] or 0
    
    pago_total = ContaPagar.objects.filter(pago=True).aggregate(
        total=Sum(
            Case(
                When(valor_pago__isnull=True, then='valor'),
                default='valor_pago',
                output_field=DecimalField()
            )
        )
    )['total'] or 0
    
    pendente_total = total_pagar - pago_total
    
    em_atraso = ContaPagar.objects.filter(
        pago=False,
        vencimento__lt=date.today()
    ).aggregate(total=Sum('valor'))['total'] or 0
    
    # Dias para frente (padrão 7, pode ser alterado via GET)
    dias_frente = int(request.GET.get('dias', 7))
    
    vence_proximos_dias = ContaPagar.objects.filter(
        pago=False,
        vencimento__gte=date.today(),
        vencimento__lte=date.today() + timedelta(days=dias_frente)
    ).aggregate(total=Sum('valor'))['total'] or 0
    
    # Contagens
    total_registros = ContaPagar.objects.count()
    pago_count = ContaPagar.objects.filter(pago=True).count()
    pendente_count = total_registros - pago_count
    atrasado_count = ContaPagar.objects.filter(
        pago=False,
        vencimento__lt=date.today()
    ).count()
    
    # Contas recentes
    contas_recentes = ContaPagar.objects.select_related(
        'conta', 'conta_financeira', 'metodo_pagamento', 'usuario'
    ).order_by('-criado_em')[:10]
    
    # Contas que vencem hoje
    contas_vencem_hoje = ContaPagar.objects.filter(
        vencimento=date.today()
    ).select_related('conta', 'usuario')
    
    context = {
        'titulo': 'Financeiro',
        'total_pagar': total_pagar,
        'pago_total': pago_total,
        'pendente_total': pendente_total,
        'em_atraso': em_atraso,
        'vence_proximos_dias': vence_proximos_dias,
        'dias_frente': dias_frente,
        'total_registros': total_registros,
        'pago_count': pago_count,
        'pendente_count': pendente_count,
        'atrasado_count': atrasado_count,
        'contas_recentes': contas_recentes,
        'contas_vencem_hoje': contas_vencem_hoje,
        'hide_sidebar': True,
    }
    
    return render(request, 'financeiro/dashboard.html', context)


@login_required
@permissao_menu_required('financeiro', 'view')
def contaspagar_list(request):
    """Listagem de contas a pagar com filtros e busca"""
    
    # Base queryset
    contas = ContaPagar.objects.select_related(
        'conta', 'subconta', 'fornecedor', 'conta_financeira', 'metodo_pagamento', 'usuario'
    )
    
    # Inicializar formulário de filtros
    form = FiltroContaPagarForm(request.GET)
    
    # Aplicar filtros
    if form.is_valid():
        classificacao = form.cleaned_data.get('classificacao')
        status = form.cleaned_data.get('status')
        fornecedor = form.cleaned_data.get('fornecedor')
        data_inicio = form.cleaned_data.get('data_inicio')
        data_fim = form.cleaned_data.get('data_fim')
        
        if classificacao:
            contas = contas.filter(classificacao=classificacao)
        
        if status:
            if status == 'pago':
                contas = contas.filter(pago=True)
            elif status == 'pendente':
                contas = contas.filter(pago=False, vencimento__gte=date.today())
            elif status == 'atrasado':
                contas = contas.filter(pago=False, vencimento__lt=date.today())
        
        if fornecedor:
            contas = contas.filter(fornecedor=fornecedor)
        
        if data_inicio:
            contas = contas.filter(vencimento__gte=data_inicio)
        
        if data_fim:
            contas = contas.filter(vencimento__lte=data_fim)
    
    # Ordenação
    ordem = request.GET.get('ordem', '-vencimento')
    contas = contas.order_by(ordem)
    
    # Filtro por busca (campo + termo)
    campo_busca = request.GET.get('campo_busca', '')
    termo_busca = request.GET.get('termo_busca', '')
    mensagem_busca_valor = None
    
    if campo_busca and termo_busca:
        if campo_busca == 'id':
            contas = contas.filter(id=termo_busca)
        elif campo_busca == 'fornecedor':
            contas = contas.filter(fornecedor__nome__icontains=termo_busca)
        elif campo_busca == 'descricao':
            contas = contas.filter(descricao__icontains=termo_busca)
        elif campo_busca == 'conta':
            contas = contas.filter(Q(conta__nome__icontains=termo_busca) | Q(subconta__nome__icontains=termo_busca))
        elif campo_busca == 'valor':
            termo_limpo = termo_busca.strip().replace(',', '.')
            try:
                # Verificar se tem comparador
                if termo_limpo.startswith('>='): 
                    valor = Decimal(termo_limpo[2:].strip())
                    contas = contas.filter(valor__gte=valor)
                elif termo_limpo.startswith('<='): 
                    valor = Decimal(termo_limpo[2:].strip())
                    contas = contas.filter(valor__lte=valor)
                elif termo_limpo.startswith('>'):
                    valor = Decimal(termo_limpo[1:].strip())
                    contas = contas.filter(valor__gt=valor)
                elif termo_limpo.startswith('<'):
                    valor = Decimal(termo_limpo[1:].strip())
                    contas = contas.filter(valor__lt=valor)
                elif termo_limpo.startswith('='):
                    valor = Decimal(termo_limpo[1:].strip())
                    contas = contas.filter(valor=valor)
                else:
                    # Sem comparador, busca valor exato ou aproximado
                    valor_busca = Decimal(termo_limpo)
                    contas_exatas = contas.filter(valor=valor_busca)
                    
                    if contas_exatas.exists():
                        # Encontrou valor exato
                        contas = contas_exatas
                    else:
                        # Não encontrou exato, buscar aproximados (±20%)
                        margem = valor_busca * Decimal('0.20')  # 20% de margem
                        valor_min = valor_busca - margem
                        valor_max = valor_busca + margem
                        contas = contas.filter(valor__gte=valor_min, valor__lte=valor_max)
                        
                        if contas.exists():
                            mensagem_busca_valor = f"Valor exato R$ {valor_busca} não encontrado. Exibindo valores aproximados entre R$ {valor_min:.2f} e R$ {valor_max:.2f}."
                        else:
                            mensagem_busca_valor = f"Nenhum valor encontrado próximo a R$ {valor_busca}."
            except Exception as e:
                # Se der erro, não filtra
                pass
    
    # Paginação
    pagina = request.GET.get('page', 1)
    por_pagina = request.GET.get('por_pagina', 20)
    
    paginator = Paginator(contas, por_pagina)
    
    try:
        contas_paginadas = paginator.page(pagina)
    except PageNotAnInteger:
        contas_paginadas = paginator.page(1)
    except EmptyPage:
        contas_paginadas = paginator.page(paginator.num_pages)
    
    context = {
        'titulo': 'Contas a Pagar',
        'contas': contas_paginadas,
        'form': form,
        'total': paginator.count,
        'campo_busca': campo_busca,
        'termo_busca': termo_busca,
        'mensagem_busca_valor': mensagem_busca_valor,
        'hide_sidebar': True,
    }
    
    return render(request, 'financeiro/contas_pagar.html', context)


@login_required
@permissao_menu_required('financeiro', 'add')
def contaspagar_create(request):
    """Criar nova conta a pagar"""
    
    if request.method == 'POST':
        form = ContaPagarForm(request.POST)
        if form.is_valid():
            conta = form.save(commit=False)
            conta.usuario = request.user
            # Adicionar fornecedor que foi criado/atualizado no clean()
            if 'fornecedor' in form.cleaned_data:
                conta.fornecedor = form.cleaned_data['fornecedor']
            conta.save()
            return redirect('financeiro:contaspagar_list')
    else:
        form = ContaPagarForm()
    
    context = {
        'titulo': 'Nova Conta a Pagar',
        'form': form,
        'acao': 'criar',
        'hide_sidebar': True,
    }
    
    return render(request, 'financeiro/contaspagar_form.html', context)


@login_required
@permissao_menu_required('financeiro', 'change')
def contaspagar_edit(request, pk):
    """Editar conta a pagar"""
    
    conta = get_object_or_404(ContaPagar, pk=pk)
    
    # Não permitir edição de contas já pagas
    if conta.pago:
        messages.error(request, 'Não é possível editar uma conta que já foi paga.')
        return redirect('financeiro:contaspagar_list')
    
    if request.method == 'POST':
        form = ContaPagarForm(request.POST, instance=conta)
        if form.is_valid():
            conta_atualizada = form.save(commit=False)
            # Adicionar fornecedor que foi criado/atualizado no clean()
            if 'fornecedor' in form.cleaned_data:
                conta_atualizada.fornecedor = form.cleaned_data['fornecedor']
            conta_atualizada.save()
            return redirect('financeiro:contaspagar_list')
    else:
        form = ContaPagarForm(instance=conta)
    
    context = {
        'titulo': 'Editar Conta a Pagar',
        'form': form,
        'acao': 'editar',
        'hide_sidebar': True,
    }
    
    return render(request, 'financeiro/contaspagar_form.html', context)


@login_required
@permissao_menu_required('financeiro', 'view')
def dar_baixa(request):
    """Página para dar baixa em contas a pagar"""
    
    # Listar apenas contas pendentes
    contas_pendentes = ContaPagar.objects.filter(pago=False).select_related(
        'fornecedor', 'conta', 'subconta'
    ).order_by('vencimento')
    
    context = {
        'titulo': 'Dar Baixa em Contas',
        'contas': contas_pendentes,
        'hide_sidebar': True,
    }
    
    return render(request, 'financeiro/dar_baixa.html', context)


@login_required
@permissao_menu_required('financeiro', 'change')
def dar_baixa_conta(request, pk):
    """Processar baixa de uma conta específica"""
    
    conta = get_object_or_404(ContaPagar, pk=pk)
    
    if request.method == 'POST':
        form = DarBaixaForm(request.POST, instance=conta)
        if form.is_valid():
            conta_pagar = form.save()
            
            # Se foi marcado como pago e tem conta financeira, criar movimentação
            if conta_pagar.pago and conta_pagar.conta_financeira and conta_pagar.data_pagamento:
                # Verificar se já existe movimentação para evitar duplicação
                movimentacao_existe = MovimentacaoFinanceira.objects.filter(
                    conta_pagar=conta_pagar,
                    origem='CONTA_PAGAR'
                ).exists()
                
                if not movimentacao_existe:
                    # Criar movimentação de saída
                    MovimentacaoFinanceira.objects.create(
                        conta_financeira=conta_pagar.conta_financeira,
                        data=conta_pagar.data_pagamento,
                        tipo='SAIDA',
                        valor=conta_pagar.valor_pago or conta_pagar.valor,
                        descricao=f'Pagamento - {conta_pagar.fornecedor.nome}: {conta_pagar.descricao or "Sem descrição"}',
                        origem='CONTA_PAGAR',
                        conta_pagar=conta_pagar,
                        categoria=conta_pagar.subconta or conta_pagar.conta,
                        usuario=request.user
                    )
            
            return redirect('financeiro:dar_baixa')
    else:
        form = DarBaixaForm(instance=conta)
    
    context = {
        'titulo': 'Dar Baixa',
        'form': form,
        'conta': conta,
        'hide_sidebar': True,
    }
    
    return render(request, 'financeiro/dar_baixa_form.html', context)


@login_required
@permissao_menu_required('financeiro', 'delete')
@require_http_methods(['POST'])
def contaspagar_delete(request, pk):
    """Deletar conta a pagar"""
    
    conta = get_object_or_404(ContaPagar, pk=pk)
    
    # Não permitir exclusão de contas já pagas
    if conta.pago:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': 'Não é possível excluir uma conta que já foi paga.'})
        else:
            messages.error(request, 'Não é possível excluir uma conta que já foi paga.')
            return redirect('financeiro:contaspagar_list')
    
    conta.delete()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    
    return redirect('financeiro:contaspagar_list')


@login_required
@permissao_menu_required('financeiro', 'change')
@require_http_methods(['POST'])
def contaspagar_pagar(request, pk):
    """Marcar conta a pagar como paga"""
    
    conta = get_object_or_404(ContaPagar, pk=pk)
    
    try:
        data_pagamento = request.POST.get('data_pagamento')
        valor_pago = request.POST.get('valor_pago')
        
        if not data_pagamento or not valor_pago:
            return JsonResponse({
                'success': False,
                'mensagem': 'Data e valor de pagamento são obrigatórios.'
            })
        
        from datetime import datetime
        data_pag = datetime.strptime(data_pagamento, '%Y-%m-%d').date()
        valor = float(valor_pago)
        
        # Validações
        if valor <= 0:
            return JsonResponse({
                'success': False,
                'mensagem': 'Valor deve ser maior que zero.'
            })
        
        # Calcular valor máximo permitido
        max_valor = conta.valor + (conta.juros or 0)
        if valor > max_valor:
            return JsonResponse({
                'success': False,
                'mensagem': f'Valor máximo permitido: R$ {max_valor:.2f}'
            })
        
        # Detectar se foi pago atrasado
        pago_atrasado = data_pag > conta.vencimento
        
        # Atualizar conta
        conta.data_pagamento = data_pag
        conta.valor_pago = valor
        conta.pago = True
        conta.pago_atrasado = pago_atrasado
        conta.save()
        
        return JsonResponse({
            'success': True,
            'mensagem': 'Conta marcada como paga com sucesso!'
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'mensagem': f'Erro ao processar: {str(e)}'
        })


@login_required
@permissao_menu_required('financeiro', 'change')
@require_http_methods(['POST'])
def contaspagar_desmarcar_pago(request, pk):
    """Desmarcar conta como paga"""
    
    conta = get_object_or_404(ContaPagar, pk=pk)
    
    try:
        conta.pago = False
        conta.data_pagamento = None
        conta.valor_pago = None
        conta.pago_atrasado = False
        conta.save()
        
        return JsonResponse({
            'success': True,
            'mensagem': 'Marca de pagamento removida.'
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'mensagem': f'Erro ao processar: {str(e)}'
        })


@login_required
def get_subcontas(request, conta_pai_id):
    """Retorna as subcontas de uma conta pai específica em formato JSON para AJAX"""
    try:
        # Buscar subcontas da conta pai selecionada
        subcontas = PlanoConta.objects.filter(
            pai_id=conta_pai_id,
            ativo=True
        ).values('id', 'codigo', 'nome').order_by('codigo', 'nome')
        
        # Formatar para Select2
        dados = [
            {
                'id': sc['id'],
                'text': f"{sc['codigo']} - {sc['nome']}"
            }
            for sc in subcontas
        ]
        
        return JsonResponse({'results': dados})
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def grafico_despesas_por_conta(request):
    """Retorna dados para gráfico de despesas por conta pai"""
    try:
        from datetime import datetime
        
        # Pegar parâmetros de data
        data_inicio = request.GET.get('data_inicio')
        data_fim = request.GET.get('data_fim')
        
        # Query base
        query = ContaPagar.objects.select_related('conta')
        
        # Aplicar filtros de data se fornecidos
        if data_inicio:
            data_inicio_obj = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            query = query.filter(vencimento__gte=data_inicio_obj)
        
        if data_fim:
            data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d').date()
            query = query.filter(vencimento__lte=data_fim_obj)
        
        # Agrupar por conta pai e somar valores
        
        # Buscar todas as contas pai (sem pai)
        contas_pai = PlanoConta.objects.filter(pai__isnull=True, ativo=True).order_by('codigo', 'nome')
        
        dados = []
        for conta_pai in contas_pai:
            # Somar valores das contas que têm esta conta como pai OU a própria conta pai
            total = query.filter(
                Q(conta=conta_pai) | Q(conta__pai=conta_pai)
            ).aggregate(total=Sum('valor'))['total'] or 0
            
            if total > 0:  # Só incluir contas com valores
                dados.append({
                    'conta': f"{conta_pai.codigo} - {conta_pai.nome}",
                    'total': float(total)
                })
        
        return JsonResponse({'dados': dados})
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@permissao_menu_required('financeiro', 'view')
def relatorios_financeiros(request):
    """Página de relatórios financeiros com filtros e processamento"""
    
    form = FiltroRelatorioForm(request.GET or None)
    dados_relatorio = None
    tipo_relatorio = None
    totalizadores = {}
    campo_busca = request.GET.get('campo_busca', '')
    termo_busca = request.GET.get('termo_busca', '')
    mensagem_busca_valor = None
    
    if request.GET and form.is_valid():
        tipo_relatorio = form.cleaned_data.get('tipo_relatorio')
        
        # Base queryset
        contas = ContaPagar.objects.select_related(
            'conta', 'subconta', 'fornecedor', 'conta_financeira', 'metodo_pagamento'
        )
        
        # Aplicar filtros comuns
        tipo_data = form.cleaned_data.get('tipo_data', 'vencimento')
        data_inicio = form.cleaned_data.get('data_inicio')
        data_fim = form.cleaned_data.get('data_fim')
        status = form.cleaned_data.get('status')
        fornecedor = form.cleaned_data.get('fornecedor')
        conta_financeira = form.cleaned_data.get('conta_financeira')
        plano_conta = form.cleaned_data.get('plano_conta')
        metodo_pagamento = form.cleaned_data.get('metodo_pagamento')
        classificacao = form.cleaned_data.get('classificacao')
        
        # Filtro de data
        if data_inicio:
            if tipo_data == 'vencimento':
                contas = contas.filter(vencimento__gte=data_inicio)
            elif tipo_data == 'pagamento':
                contas = contas.filter(data_pagamento__gte=data_inicio)
            else:  # criacao
                contas = contas.filter(criado_em__date__gte=data_inicio)
        
        if data_fim:
            if tipo_data == 'vencimento':
                contas = contas.filter(vencimento__lte=data_fim)
            elif tipo_data == 'pagamento':
                contas = contas.filter(data_pagamento__lte=data_fim)
            else:  # criacao
                contas = contas.filter(criado_em__date__lte=data_fim)
        
        # Filtro de status
        if status == 'pago':
            contas = contas.filter(pago=True)
        elif status == 'pendente':
            contas = contas.filter(pago=False, vencimento__gte=date.today())
        elif status == 'atrasado':
            contas = contas.filter(pago=False, vencimento__lt=date.today())
        
        # Outros filtros
        if fornecedor:
            contas = contas.filter(fornecedor=fornecedor)
        if conta_financeira:
            contas = contas.filter(conta_financeira=conta_financeira)
        if plano_conta:
            contas = contas.filter(Q(conta=plano_conta) | Q(conta__pai=plano_conta))
        if metodo_pagamento:
            contas = contas.filter(metodo_pagamento=metodo_pagamento)
        if classificacao:
            contas = contas.filter(classificacao=classificacao)
        
        # Processar relatório específico
        if tipo_relatorio == 'contas_periodo':
            dados_relatorio = contas.order_by('-vencimento')
            totalizadores = {
                'total_valor': contas.aggregate(Sum('valor'))['valor__sum'] or 0,
                'total_pago': contas.filter(pago=True).aggregate(
                    total=Sum(Case(
                        When(valor_pago__isnull=True, then='valor'),
                        default='valor_pago',
                        output_field=DecimalField()
                    ))
                )['total'] or 0,
                'total_pendente': contas.filter(pago=False).aggregate(Sum('valor'))['valor__sum'] or 0,
                'quantidade': contas.count()
            }
        
        elif tipo_relatorio == 'pagas_pendentes':
            dados_relatorio = {
                'pagas': contas.filter(pago=True).order_by('-data_pagamento'),
                'pendentes': contas.filter(pago=False).order_by('vencimento')
            }
            totalizadores = {
                'total_pagas': contas.filter(pago=True).aggregate(
                    total=Sum(Case(
                        When(valor_pago__isnull=True, then='valor'),
                        default='valor_pago',
                        output_field=DecimalField()
                    ))
                )['total'] or 0,
                'total_pendentes': contas.filter(pago=False).aggregate(Sum('valor'))['valor__sum'] or 0,
                'qtd_pagas': contas.filter(pago=True).count(),
                'qtd_pendentes': contas.filter(pago=False).count()
            }
        
        elif tipo_relatorio == 'em_atraso':
            dados_relatorio = contas.filter(pago=False, vencimento__lt=date.today()).order_by('vencimento')
            totalizadores = {
                'total_atrasado': dados_relatorio.aggregate(Sum('valor'))['valor__sum'] or 0,
                'quantidade': dados_relatorio.count()
            }
        
        elif tipo_relatorio == 'a_vencer':
            dias = form.cleaned_data.get('dias_vencer', 30)
            dados_relatorio = contas.filter(
                pago=False,
                vencimento__gte=date.today(),
                vencimento__lte=date.today() + timedelta(days=dias)
            ).order_by('vencimento')
            totalizadores = {
                'total_a_vencer': dados_relatorio.aggregate(Sum('valor'))['valor__sum'] or 0,
                'quantidade': dados_relatorio.count(),
                'dias': dias
            }
        
        elif tipo_relatorio == 'por_fornecedor':
            dados_relatorio = contas.values(
                'fornecedor__id',
                'fornecedor__nome'
            ).annotate(
                total_valor=Sum('valor'),
                total_pago=Sum(
                    Case(
                        When(pago=True, valor_pago__isnull=True, then='valor'),
                        When(pago=True, then='valor_pago'),
                        default=0,
                        output_field=DecimalField()
                    )
                ),
                total_pendente=Sum(
                    Case(
                        When(pago=False, then='valor'),
                        default=0,
                        output_field=DecimalField()
                    )
                ),
                quantidade=Count('id')
            ).order_by('-total_valor')
            
            totalizadores = {
                'total_geral': sum(item['total_valor'] for item in dados_relatorio),
                'total_fornecedores': dados_relatorio.count()
            }
        
        elif tipo_relatorio == 'por_plano_contas':
            dados_relatorio = contas.values(
                'conta__id',
                'conta__codigo',
                'conta__nome'
            ).annotate(
                total_valor=Sum('valor'),
                total_pago=Sum(
                    Case(
                        When(pago=True, valor_pago__isnull=True, then='valor'),
                        When(pago=True, then='valor_pago'),
                        default=0,
                        output_field=DecimalField()
                    )
                ),
                quantidade=Count('id')
            ).order_by('conta__codigo')
            
            totalizadores = {
                'total_geral': sum(item['total_valor'] for item in dados_relatorio),
                'total_contas': dados_relatorio.count()
            }
        
        elif tipo_relatorio == 'por_metodo_pagamento':
            dados_relatorio = contas.filter(pago=True).values(
                'metodo_pagamento__id',
                'metodo_pagamento__nome'
            ).annotate(
                total_valor=Sum(
                    Case(
                        When(valor_pago__isnull=True, then='valor'),
                        default='valor_pago',
                        output_field=DecimalField()
                    )
                ),
                quantidade=Count('id')
            ).order_by('-total_valor')
            
            # Adicionar contas sem método de pagamento
            sem_metodo = contas.filter(pago=True, metodo_pagamento__isnull=True).aggregate(
                total=Sum(
                    Case(
                        When(valor_pago__isnull=True, then='valor'),
                        default='valor_pago',
                        output_field=DecimalField()
                    )
                ),
                qtd=Count('id')
            )
            
            totalizadores = {
                'total_geral': sum(item['total_valor'] for item in dados_relatorio) + (sem_metodo['total'] or 0),
                'sem_metodo': sem_metodo['total'] or 0,
                'qtd_sem_metodo': sem_metodo['qtd'] or 0
            }
        
        elif tipo_relatorio == 'por_instituicao':
            dados_relatorio = contas.filter(pago=True).values(
                'conta_financeira__id',
                'conta_financeira__nome',
                'conta_financeira__tipo'
            ).annotate(
                total_valor=Sum(
                    Case(
                        When(valor_pago__isnull=True, then='valor'),
                        default='valor_pago',
                        output_field=DecimalField()
                    )
                ),
                quantidade=Count('id')
            ).order_by('-total_valor')
            
            sem_conta = contas.filter(pago=True, conta_financeira__isnull=True).aggregate(
                total=Sum(
                    Case(
                        When(valor_pago__isnull=True, then='valor'),
                        default='valor_pago',
                        output_field=DecimalField()
                    )
                ),
                qtd=Count('id')
            )
            
            totalizadores = {
                'total_geral': sum(item['total_valor'] for item in dados_relatorio) + (sem_conta['total'] or 0),
                'sem_conta': sem_conta['total'] or 0,
                'qtd_sem_conta': sem_conta['qtd'] or 0
            }
        
        elif tipo_relatorio == 'fluxo_caixa':
            # Fluxo de caixa com base nas movimentações
            if not data_inicio or not data_fim:
                # Se não tiver data, usar mês atual
                data_inicio = date.today().replace(day=1)
                from calendar import monthrange
                data_fim = date.today().replace(day=monthrange(date.today().year, date.today().month)[1])
            
            movimentacoes = MovimentacaoFinanceira.objects.filter(
                data__gte=data_inicio,
                data__lte=data_fim
            ).select_related('conta_financeira')
            
            if conta_financeira:
                movimentacoes = movimentacoes.filter(conta_financeira=conta_financeira)
            
            dados_relatorio = movimentacoes.order_by('data')
            
            totalizadores = {
                'total_entradas': movimentacoes.filter(tipo='ENTRADA').aggregate(Sum('valor'))['valor__sum'] or 0,
                'total_saidas': movimentacoes.filter(tipo='SAIDA').aggregate(Sum('valor'))['valor__sum'] or 0,
                'saldo_periodo': (movimentacoes.filter(tipo='ENTRADA').aggregate(Sum('valor'))['valor__sum'] or 0) -
                                (movimentacoes.filter(tipo='SAIDA').aggregate(Sum('valor'))['valor__sum'] or 0),
                'data_inicio': data_inicio,
                'data_fim': data_fim
            }
        
        # Aplicar busca e paginação para relatórios tipo lista
        if tipo_relatorio in ['contas_periodo', 'em_atraso', 'a_vencer'] and dados_relatorio:
            # Aplicar filtro de busca
            if campo_busca and termo_busca:
                if campo_busca == 'id':
                    dados_relatorio = dados_relatorio.filter(id=termo_busca)
                elif campo_busca == 'fornecedor':
                    dados_relatorio = dados_relatorio.filter(fornecedor__nome__icontains=termo_busca)
                elif campo_busca == 'descricao':
                    dados_relatorio = dados_relatorio.filter(descricao__icontains=termo_busca)
                elif campo_busca == 'conta':
                    dados_relatorio = dados_relatorio.filter(Q(conta__nome__icontains=termo_busca) | Q(subconta__nome__icontains=termo_busca))
                elif campo_busca == 'valor':
                    termo_limpo = termo_busca.strip().replace(',', '.')
                    try:
                        # Verificar se tem comparador
                        if termo_limpo.startswith('>='): 
                            valor = Decimal(termo_limpo[2:].strip())
                            dados_relatorio = dados_relatorio.filter(valor__gte=valor)
                        elif termo_limpo.startswith('<='): 
                            valor = Decimal(termo_limpo[2:].strip())
                            dados_relatorio = dados_relatorio.filter(valor__lte=valor)
                        elif termo_limpo.startswith('>'):
                            valor = Decimal(termo_limpo[1:].strip())
                            dados_relatorio = dados_relatorio.filter(valor__gt=valor)
                        elif termo_limpo.startswith('<'):
                            valor = Decimal(termo_limpo[1:].strip())
                            dados_relatorio = dados_relatorio.filter(valor__lt=valor)
                        elif termo_limpo.startswith('='):
                            valor = Decimal(termo_limpo[1:].strip())
                            dados_relatorio = dados_relatorio.filter(valor=valor)
                        else:
                            # Sem comparador, busca valor exato ou aproximado
                            valor_busca = Decimal(termo_limpo)
                            relatorio_exato = dados_relatorio.filter(valor=valor_busca)
                            
                            if relatorio_exato.exists():
                                # Encontrou valor exato
                                dados_relatorio = relatorio_exato
                            else:
                                # Não encontrou exato, buscar aproximados (±20%)
                                margem = valor_busca * Decimal('0.20')  # 20% de margem
                                valor_min = valor_busca - margem
                                valor_max = valor_busca + margem
                                dados_relatorio = dados_relatorio.filter(valor__gte=valor_min, valor__lte=valor_max)
                                
                                if dados_relatorio.exists():
                                    mensagem_busca_valor = f"Valor exato R$ {valor_busca} não encontrado. Exibindo valores aproximados entre R$ {valor_min:.2f} e R$ {valor_max:.2f}."
                                else:
                                    mensagem_busca_valor = f"Nenhum valor encontrado próximo a R$ {valor_busca}."
                    except Exception as e:
                        # Se der erro, não filtra
                        pass
            
            # Aplicar paginação
            pagina = request.GET.get('page', 1)
            por_pagina = request.GET.get('por_pagina', 20)
            
            paginator = Paginator(dados_relatorio, por_pagina)
            
            try:
                dados_paginados = paginator.page(pagina)
            except PageNotAnInteger:
                dados_paginados = paginator.page(1)
            except EmptyPage:
                dados_paginados = paginator.page(paginator.num_pages)
            
            dados_relatorio = dados_paginados
    
    context = {
        'titulo': 'Relatórios Financeiros',
        'form': form,
        'dados_relatorio': dados_relatorio,
        'tipo_relatorio': tipo_relatorio,
        'totalizadores': totalizadores,
        'campo_busca': campo_busca,
        'termo_busca': termo_busca,
        'mensagem_busca_valor': mensagem_busca_valor,
        'hide_sidebar': True,
    }
    
    return render(request, 'financeiro/relatorios.html', context)


@login_required
@permissao_menu_required('financeiro', 'view')
def exportar_relatorio_excel(request):
    """Exportar relatório financeiro para Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    
    # Processar os mesmos filtros do relatório
    form = FiltroRelatorioForm(request.GET)
    
    if not form.is_valid():
        return redirect('financeiro:relatorios')
    
    tipo_relatorio = form.cleaned_data.get('tipo_relatorio')
    
    # Aplicar filtros base
    contas = ContaPagar.objects.select_related('fornecedor', 'conta', 'subconta', 'conta_financeira', 'metodo_pagamento')
    
    # Aplicar filtros do formulário
    tipo_data = form.cleaned_data.get('tipo_data', 'vencimento')
    data_inicio = form.cleaned_data.get('data_inicio')
    data_fim = form.cleaned_data.get('data_fim')
    status = form.cleaned_data.get('status')
    classificacao = form.cleaned_data.get('classificacao')
    fornecedor = form.cleaned_data.get('fornecedor')
    conta_financeira = form.cleaned_data.get('conta_financeira')
    plano_conta = form.cleaned_data.get('plano_conta')
    metodo_pagamento = form.cleaned_data.get('metodo_pagamento')
    
    # Aplicar filtros de data
    if data_inicio:
        if tipo_data == 'vencimento':
            contas = contas.filter(vencimento__gte=data_inicio)
        elif tipo_data == 'emissao':
            contas = contas.filter(emissao__gte=data_inicio)
        elif tipo_data == 'pagamento':
            contas = contas.filter(data_pagamento__gte=data_inicio)
    
    if data_fim:
        if tipo_data == 'vencimento':
            contas = contas.filter(vencimento__lte=data_fim)
        elif tipo_data == 'emissao':
            contas = contas.filter(emissao__lte=data_fim)
        elif tipo_data == 'pagamento':
            contas = contas.filter(data_pagamento__lte=data_fim)
    
    # Outros filtros
    if status:
        contas = contas.filter(pago=(status == 'pago'))
    if classificacao:
        contas = contas.filter(classificacao=classificacao)
    if fornecedor:
        contas = contas.filter(fornecedor=fornecedor)
    if conta_financeira:
        contas = contas.filter(conta_financeira=conta_financeira)
    if plano_conta:
        contas = contas.filter(Q(conta=plano_conta) | Q(subconta=plano_conta))
    if metodo_pagamento:
        contas = contas.filter(metodo_pagamento=metodo_pagamento)
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Estilos
    header_fill = PatternFill(start_color="22D3EE", end_color="22D3EE", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:G1')
    titulo_map = {
        'contas_periodo': 'Contas a Pagar por Período',
        'pagas_pendentes': 'Contas Pagas vs Pendentes',
        'em_atraso': 'Contas em Atraso',
        'a_vencer': 'Contas a Vencer',
        'por_fornecedor': 'Contas por Fornecedor',
        'por_plano_contas': 'Contas por Plano de Contas',
        'por_metodo_pagamento': 'Contas por Método de Pagamento',
        'por_instituicao': 'Contas por Instituição Financeira',
        'fluxo_caixa': 'Fluxo de Caixa'
    }
    ws['A1'] = titulo_map.get(tipo_relatorio, 'Relatório Financeiro')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Cabeçalhos
    row = 3
    if tipo_relatorio in ['contas_periodo', 'pagas_pendentes', 'em_atraso', 'a_vencer']:
        headers = ['Fornecedor', 'Descrição', 'Vencimento', 'Valor', 'Status', 'Plano de Contas', 'Método Pgto']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Filtrar dados
        if tipo_relatorio == 'pagas_pendentes':
            dados = contas.order_by('vencimento')
        elif tipo_relatorio == 'em_atraso':
            dados = contas.filter(pago=False, vencimento__lt=date.today()).order_by('vencimento')
        elif tipo_relatorio == 'a_vencer':
            dias = form.cleaned_data.get('dias_vencer', 30)
            dados = contas.filter(
                pago=False,
                vencimento__gte=date.today(),
                vencimento__lte=date.today() + timedelta(days=dias)
            ).order_by('vencimento')
        else:
            dados = contas.order_by('vencimento')
        
        # Dados
        row += 1
        total = 0
        for conta in dados:
            ws.cell(row=row, column=1, value=conta.fornecedor.nome if conta.fornecedor else '')
            ws.cell(row=row, column=2, value=conta.descricao or '')
            ws.cell(row=row, column=3, value=conta.vencimento.strftime('%d/%m/%Y') if conta.vencimento else '')
            ws.cell(row=row, column=4, value=float(conta.valor))
            ws.cell(row=row, column=5, value='Pago' if conta.pago else 'Pendente')
            ws.cell(row=row, column=6, value=str(conta.subconta or conta.conta or ''))
            ws.cell(row=row, column=7, value=str(conta.metodo_pagamento or ''))
            total += conta.valor
            row += 1
        
        # Total
        row += 1
        ws.cell(row=row, column=3, value='TOTAL:')
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=4, value=float(total))
        ws.cell(row=row, column=4).font = Font(bold=True)
    
    # Ajustar largura das colunas
    from openpyxl.utils import get_column_letter
    for idx, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(idx)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar resposta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'relatorio_{tipo_relatorio}_{date.today().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
@permissao_menu_required('financeiro', 'view')
def gerar_relatorio_pdf(request):
    """Gerar relatório financeiro em PDF usando ReportLab"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from django.http import HttpResponse
    
    # Processar os mesmos filtros do relatório
    form = FiltroRelatorioForm(request.GET)
    
    if not form.is_valid():
        return redirect('financeiro:relatorios')
    
    tipo_relatorio = form.cleaned_data.get('tipo_relatorio')
    
    # Aplicar filtros base
    contas = ContaPagar.objects.select_related('fornecedor', 'conta', 'subconta', 'conta_financeira', 'metodo_pagamento')
    
    # Aplicar filtros do formulário
    tipo_data = form.cleaned_data.get('tipo_data', 'vencimento')
    data_inicio = form.cleaned_data.get('data_inicio')
    data_fim = form.cleaned_data.get('data_fim')
    status = form.cleaned_data.get('status')
    classificacao = form.cleaned_data.get('classificacao')
    fornecedor = form.cleaned_data.get('fornecedor')
    conta_financeira = form.cleaned_data.get('conta_financeira')
    plano_conta = form.cleaned_data.get('plano_conta')
    metodo_pagamento = form.cleaned_data.get('metodo_pagamento')
    
    # Aplicar filtros de data
    if data_inicio:
        if tipo_data == 'vencimento':
            contas = contas.filter(vencimento__gte=data_inicio)
        elif tipo_data == 'emissao':
            contas = contas.filter(emissao__gte=data_inicio)
        elif tipo_data == 'pagamento':
            contas = contas.filter(data_pagamento__gte=data_inicio)
    
    if data_fim:
        if tipo_data == 'vencimento':
            contas = contas.filter(vencimento__lte=data_fim)
        elif tipo_data == 'emissao':
            contas = contas.filter(emissao__lte=data_fim)
        elif tipo_data == 'pagamento':
            contas = contas.filter(data_pagamento__lte=data_fim)
    
    # Outros filtros
    if status:
        contas = contas.filter(pago=(status == 'pago'))
    if classificacao:
        contas = contas.filter(classificacao=classificacao)
    if fornecedor:
        contas = contas.filter(fornecedor=fornecedor)
    if conta_financeira:
        contas = contas.filter(conta_financeira=conta_financeira)
    if plano_conta:
        contas = contas.filter(Q(conta=plano_conta) | Q(subconta=plano_conta))
    if metodo_pagamento:
        contas = contas.filter(metodo_pagamento=metodo_pagamento)
    
    # Criar resposta HTTP
    response = HttpResponse(content_type='application/pdf')
    filename = f'relatorio_{tipo_relatorio}_{date.today().strftime("%Y%m%d")}.pdf'
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    
    # Criar documento PDF
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), 
                           rightMargin=1*cm, leftMargin=1*cm,
                           topMargin=1*cm, bottomMargin=1*cm)
    
    # Estilos
    styles = getSampleStyleSheet()
    
    # Estilo para título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#0f172a'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#64748b'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    # Mapa de títulos
    titulo_map = {
        'contas_periodo': 'Contas a Pagar por Período',
        'pagas_pendentes': 'Contas Pagas vs Pendentes',
        'em_atraso': 'Contas em Atraso',
        'a_vencer': 'Contas a Vencer',
        'por_fornecedor': 'Contas por Fornecedor',
        'por_plano_contas': 'Contas por Plano de Contas',
        'por_metodo_pagamento': 'Contas por Método de Pagamento',
        'por_instituicao': 'Contas por Instituição Financeira',
        'fluxo_caixa': 'Fluxo de Caixa'
    }
    
    # Adicionar título
    titulo = titulo_map.get(tipo_relatorio, 'Relatório Financeiro')
    elements = []
    elements.append(Paragraph(titulo, title_style))
    elements.append(Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Adicionar informações do período
    if data_inicio or data_fim:
        periodo_text = "Período: "
        if data_inicio:
            periodo_text += f"de {data_inicio.strftime('%d/%m/%Y')} "
        if data_fim:
            periodo_text += f"até {data_fim.strftime('%d/%m/%Y')}"
        elements.append(Paragraph(periodo_text, subtitle_style))
        elements.append(Spacer(1, 0.3*cm))
    
    # Preparar dados da tabela
    if tipo_relatorio in ['contas_periodo', 'pagas_pendentes', 'em_atraso', 'a_vencer']:
        # Filtrar dados conforme tipo de relatório
        if tipo_relatorio == 'pagas_pendentes':
            dados = contas.order_by('vencimento')
        elif tipo_relatorio == 'em_atraso':
            dados = contas.filter(pago=False, vencimento__lt=date.today()).order_by('vencimento')
        elif tipo_relatorio == 'a_vencer':
            dias = form.cleaned_data.get('dias_vencer', 30)
            dados = contas.filter(
                pago=False,
                vencimento__gte=date.today(),
                vencimento__lte=date.today() + timedelta(days=dias)
            ).order_by('vencimento')
        else:
            dados = contas.order_by('vencimento')
        
        # Cabeçalho da tabela
        table_data = [['Fornecedor', 'Descrição', 'Vencimento', 'Valor', 'Status']]
        
        # Dados
        total = 0
        for conta in dados:
            table_data.append([
                conta.fornecedor.nome[:30] if conta.fornecedor else '-',
                (conta.descricao or '-')[:40],
                conta.vencimento.strftime('%d/%m/%Y') if conta.vencimento else '-',
                f'R$ {conta.valor:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.'),
                'Pago' if conta.pago else 'Pendente'
            ])
            total += conta.valor
        
        # Total
        table_data.append(['', '', 'TOTAL:', f'R$ {total:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.'), ''])
        
        # Criar tabela
        table = Table(table_data, colWidths=[6*cm, 8*cm, 3*cm, 3*cm, 2.5*cm])
        
        # Estilo da tabela
        table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Corpo
            ('BACKGROUND', (0, 1), (-1, -2), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1e293b')),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 1), (2, -1), 'CENTER'),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
            ('ALIGN', (4, 1), (4, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            
            # Linhas alternadas
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8fafc')]),
            
            # Bordas
            ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#cbd5e1')),
            
            # Total
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('ALIGN', (2, -1), (2, -1), 'RIGHT'),
            ('ALIGN', (3, -1), (3, -1), 'RIGHT'),
            ('TOPPADDING', (0, -1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        
        # Rodapé
        elements.append(Spacer(1, 1*cm))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#94a3b8'),
            alignment=TA_CENTER
        )
        elements.append(Paragraph(f"Total de registros: {len(dados)}", footer_style))
        
        # Para relatório pagas_pendentes, adicionar página com resumo
        if tipo_relatorio == 'pagas_pendentes':
            from reportlab.graphics.shapes import Drawing
            from reportlab.graphics.charts.piecharts import Pie
            from reportlab.lib import colors as rl_colors
            from reportlab.platypus import PageBreak
            
            # Nova página
            elements.append(PageBreak())
            
            # Título do resumo
            resumo_title_style = ParagraphStyle(
                'ResumoTitle',
                parent=styles['Title'],
                fontSize=16,
                textColor=colors.HexColor('#0f172a'),
                spaceAfter=20,
                alignment=TA_CENTER
            )
            elements.append(Paragraph("Resumo - Contas Pagas vs Pendentes", resumo_title_style))
            elements.append(Spacer(1, 0.5*cm))
            
            # Calcular totais
            total_pago = dados.filter(pago=True).aggregate(
                total=Sum(
                    Case(
                        When(valor_pago__isnull=True, then='valor'),
                        default='valor_pago',
                        output_field=DecimalField()
                    )
                )
            )['total'] or 0
            
            total_pendente = dados.filter(pago=False).aggregate(
                total=Sum('valor')
            )['total'] or 0
            
            qtd_pago = dados.filter(pago=True).count()
            qtd_pendente = dados.filter(pago=False).count()
            
            # Tabela resumo
            resumo_data = [
                ['Status', 'Quantidade', 'Valor Total', '% do Total'],
                [
                    'Pagas',
                    str(qtd_pago),
                    f'R$ {total_pago:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.'),
                    f'{(total_pago/(total_pago+total_pendente)*100) if (total_pago+total_pendente) > 0 else 0:.1f}%'
                ],
                [
                    'Pendentes',
                    str(qtd_pendente),
                    f'R$ {total_pendente:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.'),
                    f'{(total_pendente/(total_pago+total_pendente)*100) if (total_pago+total_pendente) > 0 else 0:.1f}%'
                ],
                [
                    'TOTAL',
                    str(qtd_pago + qtd_pendente),
                    f'R$ {(total_pago + total_pendente):,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.'),
                    '100.0%'
                ]
            ]
            
            resumo_table = Table(resumo_data, colWidths=[8*cm, 4*cm, 5*cm, 5*cm])
            resumo_table.setStyle(TableStyle([
                # Cabeçalho
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Corpo
                ('BACKGROUND', (0, 1), (-1, -2), colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -2), colors.HexColor('#1e293b')),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 1), (1, -1), 'CENTER'),
                ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
                ('ALIGN', (3, 1), (3, -1), 'CENTER'),
                ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -2), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8fafc')]),
                
                # Bordas
                ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#cbd5e1')),
                
                # Total
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#0f172a')),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 10),
                ('ALIGN', (1, -1), (1, -1), 'CENTER'),
                ('ALIGN', (2, -1), (2, -1), 'RIGHT'),
                ('ALIGN', (3, -1), (3, -1), 'CENTER'),
                ('TOPPADDING', (0, -1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
            ]))
            
            elements.append(resumo_table)
            elements.append(Spacer(1, 1.5*cm))
            
            # Gráfico de pizza
            if total_pago > 0 or total_pendente > 0:
                from reportlab.lib.pagesizes import landscape, A4
                page_width = landscape(A4)[0]
                drawing = Drawing(page_width - 4*cm, 280)  # Largura total da página menos margens
                
                # Pizza por valor - centralizada
                pie = Pie()
                pie.x = (page_width - 4*cm - 200) / 2  # Centralizar pizza de 200pt
                pie.y = 40
                pie.width = 200
                pie.height = 200
                pie.data = [float(total_pago), float(total_pendente)]
                pie.labels = [
                    f'Pagas: R$ {total_pago:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.'),
                    f'Pendentes: R$ {total_pendente:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')
                ]
                pie.slices[0].fillColor = rl_colors.HexColor('#64748b')  # Cinza médio
                pie.slices[1].fillColor = rl_colors.HexColor('#cbd5e1')  # Cinza claro
                pie.slices[0].strokeColor = rl_colors.HexColor('#1e293b')
                pie.slices[1].strokeColor = rl_colors.HexColor('#1e293b')
                pie.slices[0].strokeWidth = 2
                pie.slices[1].strokeWidth = 2
                pie.sideLabels = True
                pie.simpleLabels = False
                pie.slices[0].popout = 10
                pie.slices[0].fontColor = rl_colors.black
                pie.slices[1].fontColor = rl_colors.black
                pie.slices[0].fontSize = 10
                pie.slices[1].fontSize = 10
                pie.slices[0].labelRadius = 1.25
                pie.slices[1].labelRadius = 1.25
                pie.slices[0].fontName = 'Helvetica-Bold'
                pie.slices[1].fontName = 'Helvetica-Bold'
                
                drawing.add(pie)
                elements.append(drawing)
                
                # Legenda
                legenda_style = ParagraphStyle(
                    'Legenda',
                    parent=styles['Normal'],
                    fontSize=9,
                    textColor=colors.HexColor('#64748b'),
                    alignment=TA_CENTER,
                    spaceAfter=10
                )
                elements.append(Spacer(1, 0.5*cm))
                elements.append(Paragraph("Distribuição por Valor Total", legenda_style))
    
    elif tipo_relatorio == 'por_plano_contas':
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        from reportlab.lib import colors as rl_colors
        from reportlab.platypus import PageBreak
        
        # Agrupar dados primeiro por conta pai, depois por subconta
        dados_agrupados = contas.values(
            'conta__id',
            'conta__codigo',
            'conta__nome',
            'subconta__id',
            'subconta__codigo',
            'subconta__nome'
        ).annotate(
            total_valor=Sum(
                Case(
                    When(pago=True, valor_pago__isnull=False, then='valor_pago'),
                    When(pago=True, valor_pago__isnull=True, then='valor'),
                    When(pago=False, then='valor'),
                    default=0,
                    output_field=DecimalField()
                )
            ),
            quantidade=Count('id')
        ).order_by('conta__codigo', 'subconta__codigo')
        
        # Organizar dados hierarquicamente
        hierarquia = {}
        resumo_contas_pai = {}
        
        for item in dados_agrupados:
            conta_id = item['conta__id']
            conta_codigo = item['conta__codigo']
            conta_nome = item['conta__nome']
            
            if conta_id not in hierarquia:
                hierarquia[conta_id] = {
                    'codigo': conta_codigo,
                    'nome': conta_nome,
                    'total': 0,
                    'quantidade': 0,
                    'subcontas': []
                }
                resumo_contas_pai[conta_id] = {
                    'codigo': conta_codigo,
                    'nome': conta_nome,
                    'total': 0
                }
            
            if item['subconta__id']:
                # É uma subconta
                hierarquia[conta_id]['subcontas'].append({
                    'codigo': item['subconta__codigo'],
                    'nome': item['subconta__nome'],
                    'total': item['total_valor'],
                    'quantidade': item['quantidade']
                })
            else:
                # Valores diretos na conta pai
                hierarquia[conta_id]['total'] = item['total_valor']
                hierarquia[conta_id]['quantidade'] = item['quantidade']
            
            # Acumular no resumo da conta pai
            resumo_contas_pai[conta_id]['total'] += item['total_valor']
        
        # Cabeçalho da tabela
        table_data = [['Plano de Contas', 'Quantidade', 'Total']]
        
        # Variáveis para estilo
        table_styles = [
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ]
        
        # Construir tabela com hierarquia
        total_geral = 0
        linha_atual = 1
        
        for conta_id in sorted(hierarquia.keys(), key=lambda x: hierarquia[x]['codigo']):
            conta = hierarquia[conta_id]
            
            # Conta pai (em negrito)
            table_data.append([
                f"{conta['codigo']} - {conta['nome'][:50]}",
                str(conta['quantidade']) if conta['quantidade'] > 0 else '',
                f'R$ {conta["total"]:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.') if conta['total'] > 0 else ''
            ])
            
            # Estilo para conta pai (negrito)
            table_styles.append(('FONTNAME', (0, linha_atual), (-1, linha_atual), 'Helvetica-Bold'))
            table_styles.append(('TEXTCOLOR', (0, linha_atual), (-1, linha_atual), colors.HexColor('#0f172a')))
            table_styles.append(('FONTSIZE', (0, linha_atual), (-1, linha_atual), 10))
            linha_atual += 1
            
            # Subcontas (identadas)
            for subconta in conta['subcontas']:
                table_data.append([
                    f"    • {subconta['codigo']} - {subconta['nome'][:45]}",
                    str(subconta['quantidade']),
                    f'R$ {subconta["total"]:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')
                ])
                
                # Estilo para subconta (texto normal, cinza)
                table_styles.append(('TEXTCOLOR', (0, linha_atual), (-1, linha_atual), colors.HexColor('#64748b')))
                table_styles.append(('FONTSIZE', (0, linha_atual), (-1, linha_atual), 9))
                linha_atual += 1
                
                total_geral += subconta['total']
            
            if conta['total'] > 0:
                total_geral += conta['total']
        
        # Total geral
        table_data.append(['TOTAL GERAL:', '', f'R$ {total_geral:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')])
        
        # Criar tabela
        table = Table(table_data, colWidths=[15*cm, 3*cm, 4.5*cm])
        
        # Aplicar estilos gerais
        table_styles.extend([
            # Corpo
            ('BACKGROUND', (0, 1), (-1, -2), colors.white),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),
            ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            
            # Linhas alternadas (exceto total)
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8fafc')]),
            
            # Bordas
            ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#cbd5e1')),
            
            # Total geral
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('ALIGN', (0, -1), (0, -1), 'RIGHT'),
            ('ALIGN', (2, -1), (2, -1), 'RIGHT'),
            ('TOPPADDING', (0, -1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
        ])
        
        table.setStyle(TableStyle(table_styles))
        elements.append(table)
        
        # Rodapé da primeira seção
        elements.append(Spacer(1, 1*cm))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#94a3b8'),
            alignment=TA_CENTER
        )
        elements.append(Paragraph(f"Total de planos de contas: {len(hierarquia)}", footer_style))
        
        # Nova página com resumo e gráfico
        elements.append(PageBreak())
        
        # Título do resumo
        resumo_title_style = ParagraphStyle(
            'ResumoTitle',
            parent=styles['Title'],
            fontSize=16,
            textColor=colors.HexColor('#0f172a'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        elements.append(Paragraph("Resumo por Conta Pai", resumo_title_style))
        elements.append(Spacer(1, 0.5*cm))
        
        # Tabela resumo
        resumo_data = [['Conta', 'Total']]
        chart_labels = []
        chart_values = []
        
        for conta_id in sorted(resumo_contas_pai.keys(), key=lambda x: resumo_contas_pai[x]['total'], reverse=True):
            conta = resumo_contas_pai[conta_id]
            resumo_data.append([
                f"{conta['codigo']} - {conta['nome'][:40]}",
                f'R$ {conta["total"]:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')
            ])
            chart_labels.append(f"{conta['codigo']}")
            chart_values.append(float(conta['total']))
        
        # Total no resumo
        resumo_data.append(['TOTAL:', f'R$ {total_geral:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')])
        
        resumo_table = Table(resumo_data, colWidths=[15*cm, 7*cm])
        resumo_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1e293b')),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8fafc')]),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#cbd5e1')),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('ALIGN', (0, -1), (0, -1), 'RIGHT'),
            ('ALIGN', (1, -1), (1, -1), 'RIGHT'),
        ]))
        
        elements.append(resumo_table)
        elements.append(Spacer(1, 1*cm))
        
        # Gráfico de barras
        if chart_values:
            drawing = Drawing(500, 250)
            chart = VerticalBarChart()
            chart.x = 50
            chart.y = 20
            chart.width = 400
            chart.height = 180
            chart.data = [chart_values[:10]]  # Limitar a 10 contas para não poluir
            chart.categoryAxis.categoryNames = chart_labels[:10]
            chart.categoryAxis.labels.angle = 45
            chart.categoryAxis.labels.fontSize = 8
            chart.categoryAxis.labels.textAnchor = 'end'
            chart.valueAxis.valueMin = 0
            chart.valueAxis.valueMax = max(chart_values[:10]) * 1.1 if chart_values else 100
            chart.valueAxis.valueStep = max(chart_values[:10]) / 5 if chart_values else 20
            chart.valueAxis.labels.fontSize = 9
            chart.bars[0].fillColor = rl_colors.HexColor('#0ea5e9')
            chart.bars[0].strokeColor = rl_colors.HexColor('#0f172a')
            chart.bars[0].strokeWidth = 1
            
            drawing.add(chart)
            elements.append(drawing)
            
            # Legenda do gráfico
            legenda_style = ParagraphStyle(
                'Legenda',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.HexColor('#64748b'),
                alignment=TA_CENTER,
                spaceAfter=10
            )
            elements.append(Spacer(1, 0.5*cm))
            elements.append(Paragraph("Top 10 Contas por Valor Total", legenda_style))
    
    elif tipo_relatorio == 'por_metodo_pagamento':
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.charts.barcharts import HorizontalBarChart
        from reportlab.lib import colors as rl_colors
        from reportlab.platypus import PageBreak
        
        # Agrupar por método de pagamento (apenas contas pagas)
        dados_agrupados = contas.filter(pago=True).values(
            'metodo_pagamento__id',
            'metodo_pagamento__nome'
        ).annotate(
            total_valor=Sum(
                Case(
                    When(valor_pago__isnull=True, then='valor'),
                    default='valor_pago',
                    output_field=DecimalField()
                )
            ),
            quantidade=Count('id')
        ).order_by('-total_valor')
        
        # Contas sem método de pagamento
        sem_metodo = contas.filter(pago=True, metodo_pagamento__isnull=True).aggregate(
            total=Sum(
                Case(
                    When(valor_pago__isnull=True, then='valor'),
                    default='valor_pago',
                    output_field=DecimalField()
                )
            ),
            qtd=Count('id')
        )
        
        # Cabeçalho da tabela
        table_data = [['Método de Pagamento', 'Quantidade', 'Total']]
        
        # Dados agrupados
        total_geral = 0
        for item in dados_agrupados:
            metodo_nome = item['metodo_pagamento__nome'] if item['metodo_pagamento__nome'] else 'Sem método'
            table_data.append([
                metodo_nome[:50],
                str(item['quantidade']),
                f'R$ {item["total_valor"]:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')
            ])
            total_geral += item['total_valor']
        
        # Adicionar sem método se houver
        if sem_metodo['total'] and sem_metodo['total'] > 0:
            table_data.append([
                'Sem método de pagamento',
                str(sem_metodo['qtd']),
                f'R$ {sem_metodo["total"]:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')
            ])
            total_geral += sem_metodo['total']
        
        # Total
        table_data.append(['TOTAL:', str(sum(item['quantidade'] for item in dados_agrupados) + (sem_metodo['qtd'] or 0)), f'R$ {total_geral:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')])
        
        # Criar tabela
        table = Table(table_data, colWidths=[15*cm, 3*cm, 4.5*cm])
        
        # Estilo da tabela
        table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Corpo
            ('BACKGROUND', (0, 1), (-1, -2), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1e293b')),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),
            ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            
            # Linhas alternadas
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8fafc')]),
            
            # Bordas
            ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#cbd5e1')),
            
            # Total
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('ALIGN', (0, -1), (0, -1), 'RIGHT'),
            ('ALIGN', (1, -1), (1, -1), 'CENTER'),
            ('ALIGN', (2, -1), (2, -1), 'RIGHT'),
            ('TOPPADDING', (0, -1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        
        # Rodapé
        elements.append(Spacer(1, 1*cm))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#94a3b8'),
            alignment=TA_CENTER
        )
        elements.append(Paragraph(f"Total de métodos: {len(dados_agrupados) + (1 if sem_metodo['total'] else 0)}", footer_style))
        
        # Nova página com gráfico
        if len(dados_agrupados) > 0 or (sem_metodo['total'] and sem_metodo['total'] > 0):
            elements.append(PageBreak())
            
            # Título do resumo
            resumo_title_style = ParagraphStyle(
                'ResumoTitle',
                parent=styles['Title'],
                fontSize=16,
                textColor=colors.HexColor('#0f172a'),
                spaceAfter=20,
                alignment=TA_CENTER
            )
            elements.append(Paragraph("Resumo por Método de Pagamento", resumo_title_style))
            elements.append(Spacer(1, 0.5*cm))
            
            # Preparar dados do gráfico
            chart_labels = []
            chart_values = []
            
            for item in dados_agrupados[:8]:  # Top 8
                metodo_nome = item['metodo_pagamento__nome'] if item['metodo_pagamento__nome'] else 'Sem método'
                chart_labels.append(metodo_nome[:20])
                chart_values.append(float(item['total_valor']))
            
            if sem_metodo['total'] and sem_metodo['total'] > 0 and len(chart_labels) < 8:
                chart_labels.append('Sem método')
                chart_values.append(float(sem_metodo['total']))
            
            # Gráfico de barras horizontais
            from reportlab.lib.pagesizes import landscape, A4
            page_width = landscape(A4)[0]
            
            drawing = Drawing(page_width - 4*cm, 300)
            chart = HorizontalBarChart()
            chart.x = 50
            chart.y = 20
            chart.width = page_width - 8*cm
            chart.height = 250
            chart.data = [chart_values]
            chart.categoryAxis.categoryNames = chart_labels
            chart.categoryAxis.labels.fontSize = 9
            chart.valueAxis.valueMin = 0
            chart.valueAxis.valueMax = max(chart_values) * 1.1 if chart_values else 100
            chart.valueAxis.valueStep = max(chart_values) / 5 if chart_values else 20
            chart.valueAxis.labels.fontSize = 9
            chart.bars[0].fillColor = rl_colors.HexColor('#64748b')
            chart.bars[0].strokeColor = rl_colors.HexColor('#0f172a')
            chart.bars[0].strokeWidth = 1
            
            drawing.add(chart)
            elements.append(drawing)
            
            # Legenda
            legenda_style = ParagraphStyle(
                'Legenda',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.HexColor('#64748b'),
                alignment=TA_CENTER,
                spaceAfter=10
            )
            elements.append(Spacer(1, 0.5*cm))
            elements.append(Paragraph("Métodos de Pagamento por Valor Total", legenda_style))
    
    elif tipo_relatorio == 'por_instituicao':
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        from reportlab.lib import colors as rl_colors
        from reportlab.platypus import PageBreak
        
        # Agrupar por instituição financeira (apenas contas pagas)
        dados_agrupados = contas.filter(pago=True).values(
            'conta_financeira__id',
            'conta_financeira__nome',
            'conta_financeira__tipo'
        ).annotate(
            total_valor=Sum(
                Case(
                    When(valor_pago__isnull=True, then='valor'),
                    default='valor_pago',
                    output_field=DecimalField()
                )
            ),
            quantidade=Count('id')
        ).order_by('-total_valor')
        
        # Contas sem instituição financeira
        sem_instituicao = contas.filter(pago=True, conta_financeira__isnull=True).aggregate(
            total=Sum(
                Case(
                    When(valor_pago__isnull=True, then='valor'),
                    default='valor_pago',
                    output_field=DecimalField()
                )
            ),
            qtd=Count('id')
        )
        
        # Cabeçalho da tabela
        table_data = [['Instituição Financeira', 'Tipo', 'Quantidade', 'Total']]
        
        # Dados agrupados
        total_geral = 0
        for item in dados_agrupados:
            instituicao_nome = item['conta_financeira__nome'] if item['conta_financeira__nome'] else 'Sem instituição'
            tipo = item['conta_financeira__tipo'] if item['conta_financeira__tipo'] else '-'
            table_data.append([
                instituicao_nome[:40],
                tipo[:15],
                str(item['quantidade']),
                f'R$ {item["total_valor"]:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')
            ])
            total_geral += item['total_valor']
        
        # Adicionar sem instituição se houver
        if sem_instituicao['total'] and sem_instituicao['total'] > 0:
            table_data.append([
                'Sem instituição financeira',
                '-',
                str(sem_instituicao['qtd']),
                f'R$ {sem_instituicao["total"]:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')
            ])
            total_geral += sem_instituicao['total']
        
        # Total
        table_data.append(['TOTAL:', '', str(sum(item['quantidade'] for item in dados_agrupados) + (sem_instituicao['qtd'] or 0)), f'R$ {total_geral:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')])
        
        # Criar tabela
        table = Table(table_data, colWidths=[12*cm, 3*cm, 3*cm, 4.5*cm])
        
        # Estilo da tabela
        table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Corpo
            ('BACKGROUND', (0, 1), (-1, -2), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1e293b')),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),
            ('ALIGN', (2, 1), (2, -1), 'CENTER'),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            
            # Linhas alternadas
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8fafc')]),
            
            # Bordas
            ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#cbd5e1')),
            
            # Total
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('ALIGN', (0, -1), (0, -1), 'RIGHT'),
            ('ALIGN', (2, -1), (2, -1), 'CENTER'),
            ('ALIGN', (3, -1), (3, -1), 'RIGHT'),
            ('TOPPADDING', (0, -1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        
        # Rodapé
        elements.append(Spacer(1, 1*cm))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#94a3b8'),
            alignment=TA_CENTER
        )
        elements.append(Paragraph(f"Total de instituições: {len(dados_agrupados) + (1 if sem_instituicao['total'] else 0)}", footer_style))
        
        # Nova página com gráfico
        if len(dados_agrupados) > 0 or (sem_instituicao['total'] and sem_instituicao['total'] > 0):
            elements.append(PageBreak())
            
            # Título do resumo
            resumo_title_style = ParagraphStyle(
                'ResumoTitle',
                parent=styles['Title'],
                fontSize=16,
                textColor=colors.HexColor('#0f172a'),
                spaceAfter=20,
                alignment=TA_CENTER
            )
            elements.append(Paragraph("Resumo por Instituição Financeira", resumo_title_style))
            elements.append(Spacer(1, 0.5*cm))
            
            # Preparar dados do gráfico
            chart_labels = []
            chart_values = []
            
            for item in dados_agrupados[:10]:  # Top 10
                instituicao_nome = item['conta_financeira__nome'] if item['conta_financeira__nome'] else 'Sem instituição'
                chart_labels.append(instituicao_nome[:20])
                chart_values.append(float(item['total_valor']))
            
            if sem_instituicao['total'] and sem_instituicao['total'] > 0 and len(chart_labels) < 10:
                chart_labels.append('Sem instituição')
                chart_values.append(float(sem_instituicao['total']))
            
            # Gráfico de barras verticais
            from reportlab.lib.pagesizes import landscape, A4
            page_width = landscape(A4)[0]
            
            drawing = Drawing(page_width - 4*cm, 300)
            chart = VerticalBarChart()
            chart.x = 50
            chart.y = 40
            chart.width = page_width - 8*cm
            chart.height = 200
            chart.data = [chart_values]
            chart.categoryAxis.categoryNames = chart_labels
            chart.categoryAxis.labels.angle = 45
            chart.categoryAxis.labels.fontSize = 8
            chart.categoryAxis.labels.textAnchor = 'end'
            chart.valueAxis.valueMin = 0
            chart.valueAxis.valueMax = max(chart_values) * 1.1 if chart_values else 100
            chart.valueAxis.valueStep = max(chart_values) / 5 if chart_values else 20
            chart.valueAxis.labels.fontSize = 9
            chart.bars[0].fillColor = rl_colors.HexColor('#64748b')
            chart.bars[0].strokeColor = rl_colors.HexColor('#0f172a')
            chart.bars[0].strokeWidth = 1
            
            drawing.add(chart)
            elements.append(drawing)
            
            # Legenda
            legenda_style = ParagraphStyle(
                'Legenda',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.HexColor('#64748b'),
                alignment=TA_CENTER,
                spaceAfter=10
            )
            elements.append(Spacer(1, 0.5*cm))
            elements.append(Paragraph("Top 10 Instituições Financeiras por Valor Total", legenda_style))
    
    elif tipo_relatorio == 'fluxo_caixa':
        from reportlab.platypus import PageBreak
        
        # Fluxo de caixa com base nas movimentações
        if not data_inicio or not data_fim:
            # Se não tiver data, usar mês atual
            data_inicio = date.today().replace(day=1)
            from calendar import monthrange
            data_fim = date.today().replace(day=monthrange(date.today().year, date.today().month)[1])
        
        movimentacoes = MovimentacaoFinanceira.objects.filter(
            data__gte=data_inicio,
            data__lte=data_fim
        ).select_related('conta_financeira')
        
        if conta_financeira:
            movimentacoes = movimentacoes.filter(conta_financeira=conta_financeira)
        
        movimentacoes = movimentacoes.order_by('data')
        
        # Cabeçalho da tabela
        table_data = [['Data', 'Descrição', 'Conta', 'Entrada', 'Saída', 'Saldo']]
        
        # Calcular saldo inicial
        saldo_atual = 0
        if conta_financeira:
            saldo_atual = conta_financeira.saldo_inicial or 0
        
        total_entradas = 0
        total_saidas = 0
        
        # Processar movimentações
        for mov in movimentacoes:
            entrada = ''
            saida = ''
            
            if mov.tipo == 'ENTRADA':
                entrada = f'R$ {mov.valor:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')
                saldo_atual += mov.valor
                total_entradas += mov.valor
            else:  # SAIDA
                saida = f'R$ {mov.valor:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')
                saldo_atual -= mov.valor
                total_saidas += mov.valor
            
            table_data.append([
                mov.data.strftime('%d/%m/%Y'),
                (mov.descricao or '-')[:35],
                (mov.conta_financeira.nome if mov.conta_financeira else '-')[:20],
                entrada,
                saida,
                f'R$ {saldo_atual:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')
            ])
        
        # Linha de totais
        table_data.append([
            '',
            'TOTAIS:',
            '',
            f'R$ {total_entradas:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.'),
            f'R$ {total_saidas:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.'),
            f'R$ {saldo_atual:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')
        ])
        
        # Criar tabela
        table = Table(table_data, colWidths=[2.5*cm, 8*cm, 5*cm, 3*cm, 3*cm, 3*cm])
        
        # Estilo da tabela
        table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Corpo
            ('BACKGROUND', (0, 1), (-1, -2), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1e293b')),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
            ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
            ('ALIGN', (5, 1), (5, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            
            # Linhas alternadas
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8fafc')]),
            
            # Bordas
            ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#cbd5e1')),
            
            # Total
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('ALIGN', (1, -1), (1, -1), 'RIGHT'),
            ('ALIGN', (3, -1), (3, -1), 'RIGHT'),
            ('ALIGN', (4, -1), (4, -1), 'RIGHT'),
            ('ALIGN', (5, -1), (5, -1), 'RIGHT'),
            ('TOPPADDING', (0, -1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        
        # Rodapé
        elements.append(Spacer(1, 1*cm))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#94a3b8'),
            alignment=TA_CENTER
        )
        saldo_final = total_entradas - total_saidas
        elements.append(Paragraph(f"Total de movimentações: {len(movimentacoes)} | Saldo Período: R$ {saldo_final:,.2f}".replace(',', '_').replace('.', ',').replace('_', '.'), footer_style))
    
    elif tipo_relatorio == 'por_fornecedor':
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.charts.barcharts import HorizontalBarChart
        from reportlab.lib import colors as rl_colors
        from reportlab.platypus import PageBreak
        
        # Agrupar por fornecedor
        dados_agrupados = contas.values(
            'fornecedor__id',
            'fornecedor__nome'
        ).annotate(
            total_valor=Sum(
                Case(
                    When(pago=True, valor_pago__isnull=False, then='valor_pago'),
                    When(pago=True, valor_pago__isnull=True, then='valor'),
                    When(pago=False, then='valor'),
                    default=0,
                    output_field=DecimalField()
                )
            ),
            quantidade=Count('id')
        ).order_by('-total_valor')
        
        # Contas sem fornecedor
        sem_fornecedor = contas.filter(fornecedor__isnull=True).aggregate(
            total=Sum(
                Case(
                    When(pago=True, valor_pago__isnull=False, then='valor_pago'),
                    When(pago=True, valor_pago__isnull=True, then='valor'),
                    When(pago=False, then='valor'),
                    default=0,
                    output_field=DecimalField()
                )
            ),
            qtd=Count('id')
        )
        
        # Cabeçalho da tabela
        table_data = [['Fornecedor', 'Quantidade', 'Total']]
        
        # Dados agrupados
        total_geral = 0
        for item in dados_agrupados:
            fornecedor_nome = item['fornecedor__nome'] if item['fornecedor__nome'] else 'Sem fornecedor'
            table_data.append([
                fornecedor_nome[:50],
                str(item['quantidade']),
                f'R$ {item["total_valor"]:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')
            ])
            total_geral += item['total_valor']
        
        # Adicionar sem fornecedor se houver
        if sem_fornecedor['total'] and sem_fornecedor['total'] > 0:
            table_data.append([
                'Sem fornecedor',
                str(sem_fornecedor['qtd']),
                f'R$ {sem_fornecedor["total"]:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')
            ])
            total_geral += sem_fornecedor['total']
        
        # Total
        table_data.append(['TOTAL:', str(sum(item['quantidade'] for item in dados_agrupados) + (sem_fornecedor['qtd'] or 0)), f'R$ {total_geral:,.2f}'.replace(',', '_').replace('.', ',').replace('_', '.')])
        
        # Criar tabela
        table = Table(table_data, colWidths=[15*cm, 3*cm, 4.5*cm])
        
        # Estilo da tabela
        table.setStyle(TableStyle([
            # Cabeçalho
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Corpo
            ('BACKGROUND', (0, 1), (-1, -2), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1e293b')),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),
            ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            
            # Linhas alternadas
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8fafc')]),
            
            # Bordas
            ('GRID', (0, 0), (-1, -2), 0.5, colors.HexColor('#cbd5e1')),
            
            # Total
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10),
            ('ALIGN', (0, -1), (0, -1), 'RIGHT'),
            ('ALIGN', (1, -1), (1, -1), 'CENTER'),
            ('ALIGN', (2, -1), (2, -1), 'RIGHT'),
            ('TOPPADDING', (0, -1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        
        # Rodapé
        elements.append(Spacer(1, 1*cm))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#94a3b8'),
            alignment=TA_CENTER
        )
        elements.append(Paragraph(f"Total de fornecedores: {len(dados_agrupados) + (1 if sem_fornecedor['total'] else 0)}", footer_style))
        
        # Nova página com gráfico
        if len(dados_agrupados) > 0:
            elements.append(PageBreak())
            
            # Título do resumo
            resumo_title_style = ParagraphStyle(
                'ResumoTitle',
                parent=styles['Title'],
                fontSize=16,
                textColor=colors.HexColor('#0f172a'),
                spaceAfter=20,
                alignment=TA_CENTER
            )
            elements.append(Paragraph("Top 10 Fornecedores", resumo_title_style))
            elements.append(Spacer(1, 0.5*cm))
            
            # Preparar dados do gráfico
            chart_labels = []
            chart_values = []
            
            for item in dados_agrupados[:10]:  # Top 10
                fornecedor_nome = item['fornecedor__nome'] if item['fornecedor__nome'] else 'Sem fornecedor'
                chart_labels.append(fornecedor_nome[:25])
                chart_values.append(float(item['total_valor']))
            
            # Gráfico de barras horizontais
            from reportlab.lib.pagesizes import landscape, A4
            page_width = landscape(A4)[0]
            
            drawing = Drawing(page_width - 4*cm, 350)
            chart = HorizontalBarChart()
            chart.x = 50
            chart.y = 20
            chart.width = page_width - 8*cm
            chart.height = 300
            chart.data = [chart_values]
            chart.categoryAxis.categoryNames = chart_labels
            chart.categoryAxis.labels.fontSize = 9
            chart.valueAxis.valueMin = 0
            chart.valueAxis.valueMax = max(chart_values) * 1.1 if chart_values else 100
            chart.valueAxis.valueStep = max(chart_values) / 5 if chart_values else 20
            chart.valueAxis.labels.fontSize = 9
            chart.bars[0].fillColor = rl_colors.HexColor('#64748b')
            chart.bars[0].strokeColor = rl_colors.HexColor('#0f172a')
            chart.bars[0].strokeWidth = 1
            
            drawing.add(chart)
            elements.append(drawing)
            
            # Legenda
            legenda_style = ParagraphStyle(
                'Legenda',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.HexColor('#64748b'),
                alignment=TA_CENTER,
                spaceAfter=10
            )
            elements.append(Spacer(1, 0.5*cm))
            elements.append(Paragraph("Fornecedores por Valor Total", legenda_style))
    
    # Gerar PDF
    doc.build(elements)
    
    return response
