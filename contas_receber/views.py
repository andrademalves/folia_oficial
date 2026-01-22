from django.shortcuts import render, redirect, get_object_or_404
from usuarios.decorators import verificar_permissao_menu, verificar_permissao_acao
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q, Sum
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta
from decimal import Decimal
from collections import defaultdict
from calendar import month_name


from .models import Cliente, NotaFiscal, OrigemCobranca, Parcela, CreditoCobranca, NotaFiscalCalculada
from .forms import (ClienteForm, NotaFiscalForm, OrigemCobrancaForm, ParcelaForm,
                    RegistrarParcelasForm, CreditoCobrancaForm, BaixaParcelaForm)
from importacoes.models import NotaFiscalFutura, CadastroFutura
from cadastros.models import ContaFinanceira


@login_required
def dashboard(request):
    """Dashboard principal do contas a receber"""
    # Resumo de dados
    total_notas = NotaFiscal.objects.filter(ativo=True).count()
    total_parcelas = Parcela.objects.count()
    parcelas_pendentes = Parcela.objects.filter(status_pagamento='pendente').count()
    parcelas_vencidas = Parcela.objects.filter(
        status_pagamento='pendente',
        data_vencimento__lt=timezone.now().date()
    ).count()
    
    valor_total_receber = Parcela.objects.filter(
        status_pagamento__in=['pendente', 'parcial']
    ).aggregate(total=Sum('valor'))['total'] or Decimal('0.00')
    
    valor_recebido = Parcela.objects.filter(
        status_pagamento='pago'
    ).aggregate(total=Sum('valor_pago'))['total'] or Decimal('0.00')
    
    context = {
        'hide_sidebar': True,
        'total_notas': total_notas,
        'total_parcelas': total_parcelas,
        'parcelas_pendentes': parcelas_pendentes,
        'parcelas_vencidas': parcelas_vencidas,
        'valor_total_receber': valor_total_receber,
        'valor_recebido': valor_recebido,
    }
    return render(request, 'contas_receber/dashboard.html', context)


# ==================== NOTAS FISCAIS ====================

@login_required
def lista_notas_fiscais(request):
    """Lista notas fiscais importadas do Futura que ainda não tiveram parcelas lançadas"""
    from django.db.models import Q, Count

    # Buscar notas fiscais importadas que ainda não tiveram parcelas lançadas
    notas = NotaFiscalFutura.objects.filter(parcelas_lancadas=False)

    # Filtros
    numero = request.GET.get('numero', '').strip()
    serie = request.GET.get('serie', '').strip()
    cpf_cnpj = request.GET.get('cpf_cnpj', '').strip()
    cliente = request.GET.get('cliente', '').strip()
    data_inicio = request.GET.get('data_inicio', '').strip()
    data_fim = request.GET.get('data_fim', '').strip()
    valor_min = request.GET.get('valor_min', '').strip()
    valor_max = request.GET.get('valor_max', '').strip()

    if numero:
        notas = notas.filter(nro_nota__icontains=numero)
    if serie:
        notas = notas.filter(serie__icontains=serie)
    if cpf_cnpj:
        notas = notas.filter(cnpj_cpf_cliente__icontains=cpf_cnpj)
    if cliente:
        # Precisa buscar pelo nome do cliente no CadastroFutura
        clientes_ids = CadastroFutura.objects.filter(
            razao_social__icontains=cliente
        ).values_list('id', flat=True)
        notas = notas.filter(fk_cadastro__in=clientes_ids)
    if data_inicio:
        notas = notas.filter(data_emissao__gte=data_inicio)
    if data_fim:
        notas = notas.filter(data_emissao__lte=data_fim)
    if valor_min:
        notas = notas.filter(total_nota__gte=Decimal(valor_min))
    if valor_max:
        notas = notas.filter(total_nota__lte=Decimal(valor_max))

    # Paginação
    paginator = Paginator(notas, 15)
    page = request.GET.get('page')

    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    # Buscar contas financeiras para o modal
    from cadastros.models import ContaFinanceira
    contas_financeiras = ContaFinanceira.objects.filter(ativo=True)

    # Buscar todos os clientes referenciados nas notas para exibir nome/cnpj
    clientes_ids = list(notas.values_list('fk_cadastro', flat=True))
    clientes_futura = CadastroFutura.objects.filter(id__in=clientes_ids)

    context = {
        'hide_sidebar': True,
        'page_obj': page_obj,
        'contas_financeiras': contas_financeiras,
        'clientes_futura': clientes_futura,
        'filtros': {
            'numero': numero,
            'serie': serie,
            'cpf_cnpj': cpf_cnpj,
            'cliente': cliente,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'valor_min': valor_min,
            'valor_max': valor_max,
        }
    }
    return render(request, 'contas_receber/lista_notas_fiscais.html', context)


@login_required
def registrar_parcelas(request):
    """Tela para configurar parcelas e vincular instituições financeiras"""
    # Imports necessários para o fluxo completo da função
    from datetime import datetime
    from importacoes.models import ContaParcelaFutura, CadastroFutura, NotaFiscalFutura
    from django.db.models import Q
    
    if request.method == 'POST':
        # Salvar as alterações nas parcelas
        nro_nota = request.POST.get('numero_nota')
        conta_parcelas_id = request.POST.get('conta_financeira_parcelas')
        tem_carteira = request.POST.get('tem_carteira') == '1'
        conta_carteira_id = request.POST.get('conta_financeira_carteira')
        
        try:
            
            # Verificar se é nota importada ou do sistema
            try:
                nota = NotaFiscal.objects.get(numero_nota=nro_nota, ativo=True)
                modo_importada = False
            except NotaFiscal.DoesNotExist:
                # Converter para inteiro pois nro_nota é IntegerField
                nota = NotaFiscalFutura.objects.get(nro_nota=int(nro_nota))
                modo_importada = True
            
            if modo_importada:
                # Para notas importadas, criar NotaFiscal e Parcelas no sistema
                
                # 1. Criar ou buscar o cliente
                cadastro = CadastroFutura.objects.get(id=nota.fk_cadastro)
                cliente, _ = Cliente.objects.get_or_create(
                    cpf_cnpj=cadastro.cnpj_cpf,
                    defaults={
                        'nome': cadastro.razao_social or cadastro.fantasia,
                        'email': cadastro.e_mail,  # Campo é e_mail com underline
                        'ativo': True
                    }
                )
                
                # 2. Criar NotaFiscal no sistema
                nota_fiscal, _ = NotaFiscal.objects.get_or_create(
                    numero_nota=str(nota.nro_nota),
                    serie=nota.serie or '',
                    defaults={
                        'cliente': cliente,
                        'numero_pedido': str(nota.fk_pedido) if nota.fk_pedido else None,
                        'data_emissao': nota.data_emissao,
                        'valor_produtos': nota.total_produto,
                        'valor_ipi': nota.total_ipi_valor,
                        'valor_desconto': nota.total_desconto,
                        'valor_acrescimo': nota.total_acrescimo,
                        'valor_total': nota.total_nota,
                        'ativo': True
                    }
                )
                
                # 3. Buscar parcelas importadas do Futura (tentar ambos padrões)
                from django.db.models import Q
                parcelas_futura = ContaParcelaFutura.objects.filter(
                    Q(documento__contains=f'NT:{nota.nro_nota}') | 
                    Q(documento__contains=f'NF:{nota.nro_nota}')
                ).order_by('data_vencimento')
                
                # 4. Criar parcelas NF
                for idx, p in enumerate(parcelas_futura, 1):
                    data_venc = request.POST.get(f'data_vencimento_{p.id}') or p.data_vencimento
                    if isinstance(data_venc, str):
                        data_venc = datetime.strptime(data_venc, '%Y-%m-%d').date()
                    
                    instituicao_id = request.POST.get(f'instituicao_{p.id}') or conta_parcelas_id
                    
                    Parcela.objects.create(
                        nota_fiscal=nota_fiscal,
                        cliente=cliente,
                        conta_financeira_id=instituicao_id,
                        tipo_parcela='NF',
                        numero_parcela=idx,
                        codigo_identificador=f'NT:{nota.nro_nota} {idx}/{parcelas_futura.count()}',
                        valor=p.valor_parcela,
                        data_vencimento=data_venc,
                        status_pagamento='pendente'
                    )
                
                # 5. Criar parcelas de Carteira se houver
                if tem_carteira and conta_carteira_id:
                    valor_carteira = nota.total_produto - nota.total_ipi_valor
                    quantidade_parcelas = parcelas_futura.count()
                    valor_parcela_carteira = valor_carteira / quantidade_parcelas
                    
                    # Usar nosso_numero da primeira parcela
                    numero_documento = parcelas_futura.first().nosso_numero if parcelas_futura.exists() else 'DOC'
                    
                    for idx, p in enumerate(parcelas_futura, 1):
                        data_venc = request.POST.get(f'data_vencimento_carteira_{idx}') or p.data_vencimento
                        if isinstance(data_venc, str):
                            data_venc = datetime.strptime(data_venc, '%Y-%m-%d').date()
                        
                        instituicao_id = request.POST.get(f'instituicao_carteira_{idx}') or conta_carteira_id
                        
                        Parcela.objects.create(
                            nota_fiscal=nota_fiscal,
                            cliente=cliente,
                            conta_financeira_id=instituicao_id,
                            tipo_parcela='CARTEIRA',
                            numero_parcela=idx,
                            codigo_identificador=f'{numero_documento}-{idx}',
                            valor=valor_parcela_carteira,
                            data_vencimento=data_venc,
                            status_pagamento='pendente'
                        )
                
                # 6. Marcar nota como lançada
                nota.parcelas_lancadas = True
                nota.save()
                
                messages.success(request, f'Parcelas criadas com sucesso! {parcelas_futura.count()} parcelas NF' + (f' + {parcelas_futura.count()} parcelas Carteira' if tem_carteira else ''))
                return redirect('lista_notas_fiscais')
            
            # Se não for importada, continuar com o fluxo normal
            # Buscar parcelas existentes do Futura
            parcelas_futura = Parcela.objects.filter(
                nota_fiscal=nota,
                tipo_parcela='NF'
            ).order_by('numero_parcela')
            
            for parcela in parcelas_futura:
                nova_data = request.POST.get(f'data_vencimento_{parcela.id}')
                instituicao_id = request.POST.get(f'instituicao_{parcela.id}')
                
                if nova_data:
                    parcela.data_vencimento = datetime.strptime(nova_data, '%Y-%m-%d').date()
                
                # Usar instituição específica ou a geral
                if instituicao_id:
                    parcela.conta_financeira_id = instituicao_id
                else:
                    parcela.conta_financeira_id = conta_parcelas_id
                
                # Gerar código_identificador no formato NT:NUMERO PARCELA/TOTAL
                if not parcela.codigo_identificador:
                    total_parcelas = parcelas_futura.count()
                    parcela.codigo_identificador = f'NT:{nota.numero_nota} {parcela.numero_parcela}/{total_parcelas}'
                
                parcela.save()
            
            # Se houver carteira, criar parcelas de carteira
            if tem_carteira and conta_carteira_id:
                conta_carteira = ContaFinanceira.objects.get(id=conta_carteira_id)
                valor_carteira = nota.valor_carteira
                quantidade_parcelas = parcelas_futura.count()
                valor_parcela_carteira = valor_carteira / quantidade_parcelas
                
                # Limpar parcelas de carteira existentes
                Parcela.objects.filter(nota_fiscal=nota, tipo_parcela='CARTEIRA').delete()
                
                # Buscar número do pedido para a referência
                numero_pedido = nota.numero_pedido or 'PEDIDO'
                
                # Criar parcelas de carteira com datas do formulário
                for idx in range(1, quantidade_parcelas + 1):
                    data_venc_carteira = request.POST.get(f'data_vencimento_carteira_{idx}')
                    instituicao_carteira_id = request.POST.get(f'instituicao_carteira_{idx}')
                    
                    if data_venc_carteira:
                        data_obj = datetime.strptime(data_venc_carteira, '%Y-%m-%d').date()
                        
                        # Usar instituição específica ou a geral de carteira
                        inst_id = instituicao_carteira_id if instituicao_carteira_id else conta_carteira_id
                        
                        Parcela.objects.create(
                            nota_fiscal=nota,
                            cliente=nota.cliente,
                            conta_financeira_id=inst_id,
                            tipo_parcela='CARTEIRA',
                            numero_parcela=idx,
                            valor=valor_parcela_carteira,
                            data_vencimento=data_obj,
                            status_pagamento='pendente',
                            codigo_identificador=f'{numero_pedido}-{idx}'  # Formato: NUMERO_PEDIDO-NUMERO
                        )
                
                messages.success(request, 
                    f'Parcelas configuradas com sucesso! {quantidade_parcelas} parcelas de carteira criadas.')
            else:
                messages.success(request, 'Parcelas configuradas com sucesso!')
            
            return redirect('lista_notas_fiscais')
            
        except Exception as e:
            messages.error(request, f'Erro ao processar: {str(e)}')
            return redirect('lista_notas_fiscais')
    
    # GET - Mostrar formulário de edição de parcelas
    numero_nota = request.GET.get('numero_nota')
    conta_parcelas_id = request.GET.get('conta_parcelas')
    tem_carteira = request.GET.get('tem_carteira') == '1'
    conta_carteira_id = request.GET.get('conta_carteira')
    
    try:
        # Tenta buscar a nota no modelo principal
        try:
            nota = NotaFiscal.objects.get(numero_nota=numero_nota, ativo=True)
            parcelas_nf = Parcela.objects.filter(
                nota_fiscal=nota,
                tipo_parcela='NF'
            ).order_by('numero_parcela')
            if not parcelas_nf.exists():
                messages.error(request, 'Não há parcelas para esta nota fiscal.')
                return redirect('lista_notas_fiscais')
            modo_importada = False
        except NotaFiscal.DoesNotExist:
            # Se não existe, buscar em NotaFiscalFutura
            # Converter para inteiro pois nro_nota é IntegerField
            nota = NotaFiscalFutura.objects.get(nro_nota=int(numero_nota))
            modo_importada = True
            
            # Buscar parcelas reais importadas do Futura (usando imports do topo da função)
            parcelas_futura = ContaParcelaFutura.objects.filter(
                Q(documento__contains=f'NT:{nota.nro_nota}') | 
                Q(documento__contains=f'NF:{nota.nro_nota}')
            ).order_by('data_vencimento')
            
            if not parcelas_futura.exists():
                messages.error(request, f'Não há parcelas importadas para esta nota fiscal {nota.nro_nota}.')
                return redirect('lista_notas_fiscais')
            
            # Converter para lista de dicionários para exibição
            parcelas_nf = []
            for idx, p in enumerate(parcelas_futura, 1):
                parcelas_nf.append({
                    'id': p.id,
                    'numero_parcela': idx,
                    'valor': p.valor_parcela,
                    'data_vencimento': p.data_vencimento,
                    'documento': p.documento,
                    'tipo': 'NF',
                    'id_ref': f'nf_{p.id}',
                    'referencia': f'NT:{nota.nro_nota} {idx}/{parcelas_futura.count()}'
                })


        conta_parcelas = ContaFinanceira.objects.get(id=conta_parcelas_id)
        conta_carteira = None
        parcelas_carteira = []
        contas_financeiras = ContaFinanceira.objects.filter(ativo=True)

        # Cálculo correto do valor da carteira para notas importadas ou modelo principal
        if modo_importada:
            valor_carteira = getattr(nota, 'total_produto', 0) - getattr(nota, 'total_ipi_valor', 0)
        else:
            valor_carteira = nota.valor_carteira if hasattr(nota, 'valor_carteira') else (getattr(nota, 'valor_produtos', 0) - getattr(nota, 'valor_ipi', 0))

        if tem_carteira and conta_carteira_id:
            conta_carteira = ContaFinanceira.objects.get(id=conta_carteira_id)
            quantidade_parcelas = len(parcelas_nf)
            valor_parcela_carteira = valor_carteira / quantidade_parcelas if quantidade_parcelas else 0
            
            # Para carteira, usar fk_pedido da nota fiscal
            if modo_importada:
                numero_documento = str(nota.fk_pedido) if nota.fk_pedido else f'CART:{nota.nro_nota}'
            elif not modo_importada:
                numero_documento = getattr(nota, 'numero_pedido', None) or f'CART:{nota.numero_nota}'
            else:
                numero_documento = 'CART'
            
            for idx, parcela_nf in enumerate(parcelas_nf, 1):
                parcelas_carteira.append({
                    'numero_parcela': idx,
                    'valor': valor_parcela_carteira,
                    'data_vencimento': parcela_nf['data_vencimento'] if modo_importada else parcela_nf.data_vencimento,
                    'tipo': 'CARTEIRA',
                    'id_ref': f'carteira_{idx}',
                    'referencia': f'{numero_documento} {idx}/{quantidade_parcelas}'
                })

        # Cálculo do total geral
        if modo_importada:
            total_nota = getattr(nota, 'total_nota', 0)
        else:
            total_nota = getattr(nota, 'valor_total', 0)
        
        total_geral = float(total_nota) + float(valor_carteira) if tem_carteira else float(total_nota)

        context = {
            'hide_sidebar': True,
            'nota': nota,
            'parcelas_nf': parcelas_nf,
            'parcelas_carteira': parcelas_carteira,
            'conta_parcelas': conta_parcelas,
            'tem_carteira': tem_carteira,
            'conta_carteira': conta_carteira,
            'total_parcelas': len(parcelas_nf),
            'contas_financeiras': contas_financeiras,
            'modo_importada': modo_importada,
            'valor_carteira': valor_carteira,
            'total_geral': total_geral,
            'clientes_futura': CadastroFutura.objects.all() if modo_importada else [],
        }
        return render(request, 'contas_receber/registrar_parcelas.html', context)
    except Exception as e:
        messages.error(request, f'Dados inválidos: {str(e)}')
        return redirect('lista_notas_fiscais')


# ==================== PARCELAS ====================

@login_required
def lista_parcelas(request):
    """Lista todas as parcelas com filtros"""
    parcelas = Parcela.objects.select_related('nota_fiscal', 'cliente', 'origem').all()
    
    # Filtros
    nota = request.GET.get('nota', '').strip()
    cliente_id = request.GET.get('cliente', '').strip()
    status = request.GET.get('status', '').strip()
    origem_id = request.GET.get('origem', '').strip()
    data_inicio = request.GET.get('data_inicio', '').strip()
    data_fim = request.GET.get('data_fim', '').strip()
    
    if nota:
        parcelas = parcelas.filter(nota_fiscal__numero_nota__icontains=nota)
    if cliente_id:
        parcelas = parcelas.filter(cliente_id=cliente_id)
    if status:
        parcelas = parcelas.filter(status_pagamento=status)
    if origem_id:
        parcelas = parcelas.filter(origem_id=origem_id)
    if data_inicio:
        parcelas = parcelas.filter(data_vencimento__gte=data_inicio)
    if data_fim:
        parcelas = parcelas.filter(data_vencimento__lte=data_fim)
    
    # Paginação
    paginator = Paginator(parcelas, 15)
    page = request.GET.get('page')
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    context = {
        'hide_sidebar': True,
        'page_obj': page_obj,
        'clientes': Cliente.objects.filter(ativo=True),
        'origens': OrigemCobranca.objects.filter(ativo=True),
        'notas': NotaFiscal.objects.filter(ativo=True),
        'filtros': {
            'nota': nota,
            'cliente': cliente_id,
            'status': status,
            'origem': origem_id,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
        },
        'today': date.today()
    }
    return render(request, 'contas_receber/lista_parcelas.html', context)


@login_required
def parcelas_por_vencimento(request):
    """Relatório de parcelas agrupadas por mês de vencimento"""
    parcelas = Parcela.objects.select_related('nota_fiscal', 'cliente', 'origem', 'conta_financeira').all()
    
    # Filtros
    nota_id = request.GET.get('nota', '').strip()
    cliente_id = request.GET.get('cliente', '').strip()
    origem_id = request.GET.get('origem', '').strip()
    data_inicio = request.GET.get('data_inicio', '').strip()
    data_fim = request.GET.get('data_fim', '').strip()
    codigo = request.GET.get('codigo', '').strip()
    cnpj = request.GET.get('cnpj', '').strip()
    tipo_parcela = request.GET.get('tipo_parcela', '').strip()
    status = request.GET.get('status', '').strip()
    conta_financeira_id = request.GET.get('conta_financeira', '').strip()
    
    if nota_id:
        parcelas = parcelas.filter(nota_fiscal_id=nota_id)
    if cliente_id:
        parcelas = parcelas.filter(cliente_id=cliente_id)
    if origem_id:
        parcelas = parcelas.filter(origem_id=origem_id)
    if data_inicio:
        parcelas = parcelas.filter(data_vencimento__gte=data_inicio)
    if data_fim:
        parcelas = parcelas.filter(data_vencimento__lte=data_fim)
    if codigo:
        parcelas = parcelas.filter(codigo_identificador__icontains=codigo)
    if cnpj:
        parcelas = parcelas.filter(cliente__cpf_cnpj__icontains=cnpj)
    if tipo_parcela:
        parcelas = parcelas.filter(tipo_parcela=tipo_parcela)
    if status:
        parcelas = parcelas.filter(status_pagamento=status)
    if conta_financeira_id:
        parcelas = parcelas.filter(conta_financeira_id=conta_financeira_id)
    
    # Paginação
    paginator = Paginator(parcelas, 15)
    page = request.GET.get('page')
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    from cadastros.models import ContaFinanceira
    
    context = {
        'hide_sidebar': True,
        'page_obj': page_obj,
        'clientes': Cliente.objects.filter(ativo=True),
        'origens': OrigemCobranca.objects.filter(ativo=True),
        'notas': NotaFiscal.objects.filter(ativo=True),
        'contas_financeiras': ContaFinanceira.objects.filter(ativo=True),
        'filtros': {
            'nota': nota_id,
            'cliente': cliente_id,
            'origem': origem_id,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'codigo': codigo,
            'cnpj': cnpj,
            'tipo_parcela': tipo_parcela,
            'status': status,
            'conta_financeira': conta_financeira_id,
        },
        'today': date.today()
    }
    return render(request, 'contas_receber/parcelas_por_vencimento.html', context)


@login_required
def lista_parcelas_baixa(request):
    """Lista parcelas para dar baixa"""
    parcelas = Parcela.objects.filter(status_pagamento__in=['pendente', 'parcial']).select_related('nota_fiscal', 'cliente', 'origem', 'conta_financeira')
    
    # Filtros
    nota_id = request.GET.get('nota', '').strip()
    cliente_id = request.GET.get('cliente', '').strip()
    origem_id = request.GET.get('origem', '').strip()
    data_inicio = request.GET.get('data_inicio', '').strip()
    data_fim = request.GET.get('data_fim', '').strip()
    codigo = request.GET.get('codigo', '').strip()
    cnpj = request.GET.get('cnpj', '').strip()
    tipo_parcela = request.GET.get('tipo_parcela', '').strip()
    status = request.GET.get('status', '').strip()
    conta_financeira_id = request.GET.get('conta_financeira', '').strip()
    
    if nota_id:
        parcelas = parcelas.filter(nota_fiscal_id=nota_id)
    if cliente_id:
        parcelas = parcelas.filter(cliente_id=cliente_id)
    if origem_id:
        parcelas = parcelas.filter(origem_id=origem_id)
    if data_inicio:
        parcelas = parcelas.filter(data_vencimento__gte=data_inicio)
    if data_fim:
        parcelas = parcelas.filter(data_vencimento__lte=data_fim)
    if codigo:
        parcelas = parcelas.filter(codigo_identificador__icontains=codigo)
    if cnpj:
        parcelas = parcelas.filter(cliente__cpf_cnpj__icontains=cnpj)
    if tipo_parcela:
        parcelas = parcelas.filter(tipo_parcela=tipo_parcela)
    if status:
        parcelas = parcelas.filter(status_pagamento=status)
    if conta_financeira_id:
        parcelas = parcelas.filter(conta_financeira_id=conta_financeira_id)
    
    # Paginação
    paginator = Paginator(parcelas, 15)
    page = request.GET.get('page')
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    from cadastros.models import ContaFinanceira
    
    context = {
        'hide_sidebar': True,
        'parcelas': page_obj,
        'page_obj': page_obj,
        'clientes': Cliente.objects.filter(ativo=True),
        'origens': OrigemCobranca.objects.filter(ativo=True),
        'notas': NotaFiscal.objects.filter(ativo=True),
        'contas_financeiras': ContaFinanceira.objects.filter(ativo=True),
        'filtros': {
            'nota': nota_id,
            'cliente': cliente_id,
            'origem': origem_id,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'codigo': codigo,
            'cnpj': cnpj,
            'tipo_parcela': tipo_parcela,
            'status': status,
            'conta_financeira': conta_financeira_id,
        }
    }
    return render(request, 'contas_receber/lista_parcelas_baixa.html', context)


@login_required
def dar_baixa_parcela(request, parcela_id):
    """Dar baixa em uma parcela específica"""
    parcela = get_object_or_404(Parcela, id=parcela_id)
    
    if request.method == 'POST':
        form = BaixaParcelaForm(request.POST)
        if form.is_valid():
            parcela.data_pagamento = form.cleaned_data['data_pagamento']
            parcela.valor_pago = form.cleaned_data['valor_pago']
            parcela.desconto_concedido = form.cleaned_data['desconto'] or 0
            parcela.juros = form.cleaned_data['juros'] or 0
            parcela.multa = form.cleaned_data['multa'] or 0
            parcela.motivo_desconto = form.cleaned_data['motivo_desconto']
            parcela.observacao = form.cleaned_data['observacao']
            
            # Verificar se foi pago totalmente ou parcialmente
            if parcela.valor_pago >= parcela.total_a_pagar:
                parcela.status_pagamento = 'pago'
            else:
                parcela.status_pagamento = 'parcial'
            
            parcela.save()
            
            # Criar movimentação financeira na conta corrente
            if parcela.conta_financeira:
                from financeiro.models import MovimentacaoFinanceira
                
                # Descrição da movimentação
                descricao = f"Recebimento - {parcela.codigo_identificador or 'Parcela'}"
                if parcela.cliente:
                    descricao += f" - {parcela.cliente.nome}"
                
                # Valor total recebido (com juros e multa, menos desconto)
                valor_total = parcela.valor_pago + parcela.juros + parcela.multa - parcela.desconto_concedido
                
                MovimentacaoFinanceira.objects.create(
                    conta_financeira=parcela.conta_financeira,
                    data=parcela.data_pagamento,
                    tipo='ENTRADA',
                    valor=valor_total,
                    descricao=descricao,
                    origem='CONTA_RECEBER',
                    observacoes=parcela.observacao,
                    usuario=request.user
                )
            
            messages.success(request, 'Baixa registrada com sucesso e movimentação criada na conta corrente!')
            return redirect('lista_parcelas_baixa')
    else:
        form = BaixaParcelaForm()
    
    # Buscar todas as contas financeiras para o select do modal
    from cadastros.models import ContaFinanceira
    contas_financeiras = ContaFinanceira.objects.filter(ativo=True).order_by('nome')
    
    context = {
        'hide_sidebar': True,
        'parcela': parcela,
        'form': form,
        'contas_financeiras': contas_financeiras
    }
    return render(request, 'contas_receber/dar_baixa_parcela.html', context)


# ==================== CRÉDITOS ====================

@login_required
@login_required
def lista_creditos(request):
    """Lista créditos de cobrança separados por status"""
    # Filtros
    nota_id = request.GET.get('nota', '').strip()
    cliente_id = request.GET.get('cliente', '').strip()
    status_filtro = request.GET.get('status', '').strip()
    
    creditos = CreditoCobranca.objects.select_related('nota_fiscal', 'cliente').all()
    
    if nota_id:
        creditos = creditos.filter(nota_fiscal_id=nota_id)
    if cliente_id:
        creditos = creditos.filter(cliente_id=cliente_id)
    if status_filtro:
        creditos = creditos.filter(status=status_filtro)
    
    # Separar por status para melhor visualização
    creditos_solicitados = creditos.filter(status='solicitado')
    creditos_aprovados = creditos.filter(status='aprovado')
    creditos_rejeitados = creditos.filter(status='rejeitado')
    creditos_utilizados = creditos.filter(status='utilizado')
    
    context = {
        'hide_sidebar': True,
        'creditos_solicitados': creditos_solicitados,
        'creditos_aprovados': creditos_aprovados,
        'creditos_rejeitados': creditos_rejeitados,
        'creditos_utilizados': creditos_utilizados,
        'clientes': Cliente.objects.filter(ativo=True),
        'notas': NotaFiscal.objects.filter(ativo=True),
        'filtros': {
            'nota': nota_id,
            'cliente': cliente_id,
            'status': status_filtro,
        }
    }
    return render(request, 'contas_receber/lista_creditos.html', context)


@login_required
def criar_credito(request):
    """Criar nova solicitação de crédito"""
    if request.method == 'POST':
        form = CreditoCobrancaForm(request.POST)
        if form.is_valid():
            credito = form.save(commit=False)
            credito.usuario_solicitante = request.user.username
            credito.save()
            messages.success(request, 'Crédito solicitado com sucesso!')
            return redirect('lista_creditos')
    else:
        form = CreditoCobrancaForm()
    
    context = {'hide_sidebar': True, 'form': form}
    return render(request, 'contas_receber/criar_credito.html', context)


@login_required
def api_dados_nota(request, nota_id):
    """API para retornar dados da nota fiscal em JSON"""
    try:
        nota = NotaFiscal.objects.select_related('cliente').get(id=nota_id, ativo=True)
        
        # Calcular totais e contagem de parcelas NF
        parcelas_nf = Parcela.objects.filter(
            nota_fiscal=nota,
            tipo_parcela='NF'
        )
        total_parcelas_nf = parcelas_nf.aggregate(total=Sum('valor'))['total'] or 0
        qtd_parcelas_nf = parcelas_nf.count()
        
        # Calcular totais e contagem de parcelas Carteira
        parcelas_carteira = Parcela.objects.filter(
            nota_fiscal=nota,
            tipo_parcela='CARTEIRA'
        )
        total_parcelas_carteira = parcelas_carteira.aggregate(total=Sum('valor'))['total'] or 0
        qtd_parcelas_carteira = parcelas_carteira.count()
        
        data = {
            'success': True,
            'cliente_id': nota.cliente.id,
            'cliente_nome': nota.cliente.nome,
            'cliente_cpf_cnpj': nota.cliente.cpf_cnpj,
            'qtd_parcelas_nf': qtd_parcelas_nf,
            'total_nf': float(total_parcelas_nf),
            'qtd_parcelas_carteira': qtd_parcelas_carteira,
            'total_carteira': float(total_parcelas_carteira),
            'total_geral': float(total_parcelas_nf + total_parcelas_carteira),
            'valor_produtos': float(nota.valor_produtos) if nota.valor_produtos else 0,
            'valor_total_nota': float(nota.valor_total) if nota.valor_total else 0,
        }
        return JsonResponse(data)
    except NotaFiscal.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Nota fiscal não encontrada'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return render(request, 'contas_receber/criar_credito.html', context)


@login_required
@verificar_permissao_menu('/contas-receber/aprovacoes/')
def detalhe_credito_aprovacao(request, credito_id):
    """Visualizar e aprovar crédito com valor e justificativa"""
    credito = get_object_or_404(CreditoCobranca, id=credito_id)
    
    if request.method == 'POST':
        acao = request.POST.get('acao')
        
        if acao == 'aprovar':
            valor_aprovado = Decimal(request.POST.get('valor_aprovado', 0))
            justificativa_aprovacao = request.POST.get('justificativa_aprovacao', '')
            
            if valor_aprovado <= 0:
                messages.error(request, 'Valor aprovado deve ser maior que zero.')
                return redirect('detalhe_credito_aprovacao', credito_id=credito_id)
            
            if valor_aprovado > credito.valor_credito:
                messages.warning(request, f'Valor aprovado (R$ {valor_aprovado}) é maior que o solicitado (R$ {credito.valor_credito}).')
            
            credito.status = 'aprovado'
            credito.valor_aprovado = valor_aprovado
            credito.justificativa_aprovacao = justificativa_aprovacao
            credito.data_liberacao = timezone.now()
            credito.usuario_aprovador = request.user.username
            credito.save()
            
            messages.success(request, f'Crédito aprovado: R$ {valor_aprovado}')
            return redirect('aprovacoes_creditos')
        
        elif acao == 'rejeitar':
            motivo = request.POST.get('motivo_rejeicao', '')
            credito.status = 'rejeitado'
            credito.justificativa_aprovacao = f"Rejeitado: {motivo}"
            credito.observacoes = f"Rejeitado por {request.user.username}: {motivo}"
            credito.save()
            messages.warning(request, 'Crédito rejeitado.')
            return redirect('aprovacoes_creditos')
    
    context = {
        'credito': credito,
        'hide_sidebar': True,
    }
    return render(request, 'contas_receber/detalhe_credito_aprovacao.html', context)


@login_required
@verificar_permissao_menu('/contas-receber/aprovacoes/')
def aprovacoes_creditos(request):
    """Módulo de aprovações - listar créditos pendentes de aprovação"""
    # Listar apenas créditos pendentes
    creditos_pendentes = CreditoCobranca.objects.filter(
        status='solicitado'
    ).select_related('nota_fiscal', 'cliente').order_by('-data_solicitacao')
    
    context = {
        'creditos_pendentes': creditos_pendentes,
        'hide_sidebar': True,
    }
    return render(request, 'contas_receber/aprovacoes_creditos.html', context)


@login_required
def aprovar_credito(request, credito_id):
    """Aprovar um crédito solicitado - apenas Diretoria e Administrativo"""
    credito = get_object_or_404(CreditoCobranca, id=credito_id)
    
    # Verificar se usuário tem permissão
    if not (request.user.is_superuser or request.user.groups.filter(name__in=['Diretoria', 'Administrativo']).exists()):
        messages.error(request, 'Você não tem permissão para aprovar créditos.')
        return redirect('lista_creditos')
    
    # Verificar se o usuário não está tentando aprovar seu próprio crédito
    if credito.usuario_solicitante == request.user.username:
        messages.error(request, 'Você não pode aprovar seu próprio crédito. Solicite a outro usuário autorizado.')
        return redirect('aprovacoes_creditos')
    
    if credito.status == 'solicitado':
        credito.status = 'aprovado'
        credito.data_liberacao = timezone.now()
        credito.usuario_aprovador = request.user.username
        credito.save()
        messages.success(request, f'Crédito de R$ {credito.valor_credito} aprovado com sucesso!')
    else:
        messages.warning(request, 'Este crédito já foi processado.')
    
    return redirect('aprovacoes_creditos')


@login_required
def rejeitar_credito(request, credito_id):
    """Rejeitar um crédito solicitado - apenas Diretoria e Administrativo"""
    credito = get_object_or_404(CreditoCobranca, id=credito_id)
    
    # Verificar se usuário tem permissão
    if not (request.user.is_superuser or request.user.groups.filter(name__in=['Diretoria', 'Administrativo']).exists()):
        messages.error(request, 'Você não tem permissão para rejeitar créditos.')
        return redirect('aprovacoes_creditos')
    
    if credito.status == 'solicitado':
        if request.method == 'POST':
            motivo = request.POST.get('motivo_rejeicao', '')
            credito.status = 'rejeitado'
            credito.observacoes = f"Rejeitado por {request.user.username}: {motivo}"
            credito.save()
            messages.warning(request, 'Crédito rejeitado.')
            return redirect('aprovacoes_creditos')
    else:
        messages.warning(request, 'Este crédito já foi processado.')
        return redirect('aprovacoes_creditos')
    
    return redirect('aprovacoes_creditos')


@login_required
def aplicar_credito(request, credito_id):
    """Aplicar crédito aprovado em uma parcela"""
    credito = get_object_or_404(CreditoCobranca, id=credito_id)
    
    if credito.status != 'aprovado':
        messages.error(request, 'Este crédito não está aprovado.')
        return redirect('lista_creditos')
    
    if credito.saldo_disponivel <= 0:
        messages.error(request, 'Este crédito não possui saldo disponível.')
        return redirect('lista_creditos')
    
    if request.method == 'POST':
        parcela_id = request.POST.get('parcela_id')
        valor_aplicar = Decimal(request.POST.get('valor_aplicar', 0))
        
        parcela = get_object_or_404(Parcela, id=parcela_id)
        
        # Validações
        if valor_aplicar <= 0:
            messages.error(request, 'Valor inválido.')
            return redirect('aplicar_credito', credito_id=credito_id)
        
        if valor_aplicar > credito.saldo_disponivel:
            messages.error(request, f'Valor maior que o saldo disponível (R$ {credito.saldo_disponivel}).')
            return redirect('aplicar_credito', credito_id=credito_id)
        
        if valor_aplicar > parcela.valor:
            messages.error(request, f'Valor maior que o valor da parcela (R$ {parcela.valor}).')
            return redirect('aplicar_credito', credito_id=credito_id)
        
        # Aplicar crédito
        parcela.valor -= valor_aplicar
        parcela.observacao = (parcela.observacao or '') + f"\nCrédito aplicado: R$ {valor_aplicar} em {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        parcela.save()
        
        # Atualizar crédito
        credito.valor_utilizado += valor_aplicar
        if credito.valor_utilizado >= credito.valor_aprovado:
            credito.status = 'utilizado'
            credito.data_utilizacao = timezone.now()
        credito.save()
        
        messages.success(request, f'Crédito de R$ {valor_aplicar} aplicado na parcela com sucesso!')
        return redirect('lista_creditos')
    
    # GET - Mostrar formulário
    # Buscar parcelas do cliente que ainda estão pendentes
    parcelas = Parcela.objects.filter(
        cliente=credito.cliente,
        status_pagamento='pendente'
    ).select_related('nota_fiscal')
    
    context = {
        'hide_sidebar': True,
        'credito': credito,
        'parcelas': parcelas
    }
    return render(request, 'contas_receber/aplicar_credito.html', context)


@login_required
def liberar_credito(request, credito_id):
    """Manter para compatibilidade - redireciona para aprovar"""
    return aprovar_credito(request, credito_id)


@login_required
def relatorio_creditos(request):
    """Relatório de controle de créditos aplicados"""
    # Filtros
    status = request.GET.get('status', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    cliente_id = request.GET.get('cliente', '')
    nota_fiscal_id = request.GET.get('nota_fiscal', '')
    
    creditos = CreditoCobranca.objects.select_related(
        'cliente', 'nota_fiscal'
    ).exclude(status='solicitado').order_by('-data_solicitacao')
    
    if status:
        creditos = creditos.filter(status=status)
    if data_inicio:
        creditos = creditos.filter(data_solicitacao__gte=data_inicio)
    if data_fim:
        creditos = creditos.filter(data_solicitacao__lte=data_fim)
    if cliente_id:
        creditos = creditos.filter(cliente_id=cliente_id)
    if nota_fiscal_id:
        creditos = creditos.filter(nota_fiscal_id=nota_fiscal_id)
    
    # Para cada crédito, buscar onde foi aplicado
    creditos_detalhados = []
    for credito in creditos:
        # Buscar parcelas que mencionam este crédito nas observações
        # (isso é uma aproximação - idealmente teria uma tabela de relacionamento)
        parcelas_aplicadas = Parcela.objects.filter(
            cliente=credito.cliente,
            observacao__icontains=f'Crédito aplicado'
        ).select_related('nota_fiscal')
        
        creditos_detalhados.append({
            'credito': credito,
            'parcelas': parcelas_aplicadas
        })
    
    context = {
        'hide_sidebar': True,
        'creditos_detalhados': creditos_detalhados,
        'clientes': Cliente.objects.filter(ativo=True),
        'notas_fiscais': NotaFiscal.objects.filter(ativo=True).order_by('-numero_nota'),
        'filtros': {
            'status': status,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'cliente': cliente_id,
            'nota_fiscal': nota_fiscal_id,
        }
    }
    return render(request, 'contas_receber/relatorio_creditos.html', context)


@login_required
def excluir_credito(request, credito_id):
    """Excluir um crédito"""
    credito = get_object_or_404(CreditoCobranca, id=credito_id)
    
    if request.method == 'POST':
        credito.delete()
        messages.success(request, 'Crédito excluído com sucesso!')
        return redirect('lista_creditos')
    
    return redirect('lista_creditos')


# ==================== ORIGENS DE COBRANÇA ====================

@login_required
def lista_origens(request):
    """Lista origens de cobrança"""
    origens = OrigemCobranca.objects.all()
    context = {'hide_sidebar': True, 'origens': origens}
    return render(request, 'contas_receber/lista_origens.html', context)


@login_required
def criar_origem(request):
    """Criar nova origem de cobrança"""
    if request.method == 'POST':
        form = OrigemCobrancaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Origem criada com sucesso!')
            return redirect('lista_origens')
    else:
        form = OrigemCobrancaForm()
    
    context = {'hide_sidebar': True, 'form': form}
    return render(request, 'contas_receber/form_origem.html', context)


@login_required
def editar_origem(request, origem_id):
    """Editar origem de cobrança"""
    origem = get_object_or_404(OrigemCobranca, id=origem_id)
    
    if request.method == 'POST':
        form = OrigemCobrancaForm(request.POST, instance=origem)
        if form.is_valid():
            form.save()
            messages.success(request, 'Origem atualizada com sucesso!')
            return redirect('lista_origens')
    else:
        form = OrigemCobrancaForm(instance=origem)
    
    context = {'hide_sidebar': True, 'form': form, 'origem': origem}
    return render(request, 'contas_receber/form_origem.html', context)


@login_required
def excluir_origem(request, origem_id):
    """Excluir origem de cobrança"""
    origem = get_object_or_404(OrigemCobranca, id=origem_id)
    
    if request.method == 'POST':
        origem.delete()
        messages.success(request, 'Origem excluída com sucesso!')
        return redirect('lista_origens')
    
    return redirect('lista_origens')


# ==================== EXPORTAÇÕES ====================

@login_required
def exportar_parcelas_excel(request):
    """Exportar parcelas para Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime
    
    # Aplicar mesmos filtros da lista
    parcelas = Parcela.objects.select_related('nota_fiscal', 'cliente', 'origem', 'conta_financeira').all()
    
    # Filtros
    nota_id = request.GET.get('nota', '').strip()
    cliente_id = request.GET.get('cliente', '').strip()
    origem_id = request.GET.get('origem', '').strip()
    data_inicio = request.GET.get('data_inicio', '').strip()
    data_fim = request.GET.get('data_fim', '').strip()
    codigo = request.GET.get('codigo', '').strip()
    cnpj = request.GET.get('cnpj', '').strip()
    tipo_parcela = request.GET.get('tipo_parcela', '').strip()
    status = request.GET.get('status', '').strip()
    conta_financeira_id = request.GET.get('conta_financeira', '').strip()
    
    if nota_id:
        parcelas = parcelas.filter(nota_fiscal_id=nota_id)
    if cliente_id:
        parcelas = parcelas.filter(cliente_id=cliente_id)
    if origem_id:
        parcelas = parcelas.filter(origem_id=origem_id)
    if data_inicio:
        parcelas = parcelas.filter(data_vencimento__gte=data_inicio)
    if data_fim:
        parcelas = parcelas.filter(data_vencimento__lte=data_fim)
    if codigo:
        parcelas = parcelas.filter(codigo_identificador__icontains=codigo)
    if cnpj:
        parcelas = parcelas.filter(cliente__cpf_cnpj__icontains=cnpj)
    if tipo_parcela:
        parcelas = parcelas.filter(tipo_parcela=tipo_parcela)
    if status:
        parcelas = parcelas.filter(status_pagamento=status)
    if conta_financeira_id:
        parcelas = parcelas.filter(conta_financeira_id=conta_financeira_id)
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parcelas por Vencimento"
    
    # Cabeçalho do relatório
    ws.merge_cells('A1:J1')
    titulo_cell = ws['A1']
    titulo_cell.value = 'RELATÓRIO DE PARCELAS POR VENCIMENTO'
    titulo_cell.font = Font(size=16, bold=True, color='FFFFFF')
    titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
    titulo_cell.fill = PatternFill(start_color='0ea5e9', end_color='0ea5e9', fill_type='solid')
    ws.row_dimensions[1].height = 30
    
    # Data e hora da exportação
    ws.merge_cells('A2:J2')
    data_hora_cell = ws['A2']
    data_hora_cell.value = f'Exportado em: {datetime.now().strftime("%d/%m/%Y às %H:%M:%S")}'
    data_hora_cell.font = Font(size=10, italic=True)
    data_hora_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 20
    
    # Linha em branco
    ws.append([])
    
    # Cabeçalhos das colunas
    headers = ['Código/Ref', 'NF', 'Cliente', 'CNPJ/CPF', 'Tipo', 'Parcela', 'Valor', 'Vencimento', 'Instituição', 'Status']
    ws.append(headers)
    
    # Estilizar cabeçalho das colunas
    header_fill = PatternFill(start_color='0f172a', end_color='0f172a', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color='FFFFFF'),
        bottom=Side(style='thin', color='FFFFFF')
    )
    
    for cell in ws[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Dados
    for p in parcelas:
        ws.append([
            p.codigo_identificador,
            p.nota_fiscal.numero_nota,
            p.cliente.nome,
            p.cliente.cpf_cnpj,
            p.tipo_parcela,
            p.numero_parcela,
            float(p.valor),
            p.data_vencimento.strftime('%d/%m/%Y'),
            p.conta_financeira.nome if p.conta_financeira else '-',
            p.get_status_pagamento_display()
        ])
    
    # Ajustar largura das colunas
    column_widths = [15, 12, 30, 18, 10, 10, 12, 15, 20, 12]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Preparar response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="parcelas_vencimento_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    
    return response


@login_required
def exportar_parcelas_pdf(request):
    """Exportar parcelas para PDF"""
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from datetime import datetime
    
    # Aplicar mesmos filtros
    parcelas = Parcela.objects.select_related('nota_fiscal', 'cliente', 'origem', 'conta_financeira').all()
    
    # Filtros
    nota_id = request.GET.get('nota', '').strip()
    cliente_id = request.GET.get('cliente', '').strip()
    origem_id = request.GET.get('origem', '').strip()
    data_inicio = request.GET.get('data_inicio', '').strip()
    data_fim = request.GET.get('data_fim', '').strip()
    codigo = request.GET.get('codigo', '').strip()
    cnpj = request.GET.get('cnpj', '').strip()
    tipo_parcela = request.GET.get('tipo_parcela', '').strip()
    status = request.GET.get('status', '').strip()
    conta_financeira_id = request.GET.get('conta_financeira', '').strip()
    
    if nota_id:
        parcelas = parcelas.filter(nota_fiscal_id=nota_id)
    if cliente_id:
        parcelas = parcelas.filter(cliente_id=cliente_id)
    if origem_id:
        parcelas = parcelas.filter(origem_id=origem_id)
    if data_inicio:
        parcelas = parcelas.filter(data_vencimento__gte=data_inicio)
    if data_fim:
        parcelas = parcelas.filter(data_vencimento__lte=data_fim)
    if codigo:
        parcelas = parcelas.filter(codigo_identificador__icontains=codigo)
    if cnpj:
        parcelas = parcelas.filter(cliente__cpf_cnpj__icontains=cnpj)
    if tipo_parcela:
        parcelas = parcelas.filter(tipo_parcela=tipo_parcela)
    if status:
        parcelas = parcelas.filter(status_pagamento=status)
    if conta_financeira_id:
        parcelas = parcelas.filter(conta_financeira_id=conta_financeira_id)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="parcelas_vencimento_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    p = canvas.Canvas(response, pagesize=landscape(A4))
    width, height = landscape(A4)
    
    def draw_header(p, width, height):
        # Retângulo de fundo do título
        p.setFillColorRGB(0.055, 0.647, 0.914)  # #0ea5e9
        p.rect(1*cm, height - 2.5*cm, width - 2*cm, 1*cm, fill=True, stroke=False)
        
        # Título
        p.setFillColorRGB(1, 1, 1)  # Branco
        p.setFont("Helvetica-Bold", 16)
        p.drawCentredString(width/2, height - 2.1*cm, "RELATÓRIO DE PARCELAS POR VENCIMENTO")
        
        # Data e hora da exportação
        p.setFillColorRGB(0, 0, 0)  # Preto
        p.setFont("Helvetica-Oblique", 9)
        p.drawCentredString(width/2, height - 3*cm, f'Exportado em: {datetime.now().strftime("%d/%m/%Y às %H:%M:%S")}')
    
    # Desenhar cabeçalho
    draw_header(p, width, height)
    
    # Cabeçalhos da tabela
    y = height - 4*cm
    p.setFillColorRGB(0.059, 0.090, 0.165)  # #0f172a
    p.rect(0.5*cm, y - 0.3*cm, width - 1*cm, 0.6*cm, fill=True, stroke=False)
    
    p.setFillColorRGB(1, 1, 1)  # Branco
    p.setFont("Helvetica-Bold", 8)
    p.drawString(0.7*cm, y, "Código")
    p.drawString(3*cm, y, "NF")
    p.drawString(5*cm, y, "Cliente")
    p.drawString(10*cm, y, "CNPJ/CPF")
    p.drawString(14.5*cm, y, "Tipo")
    p.drawString(16.5*cm, y, "Parc")
    p.drawString(18*cm, y, "Valor")
    p.drawString(21*cm, y, "Vencimento")
    p.drawString(24.5*cm, y, "Instituição")
    
    # Dados
    p.setFillColorRGB(0, 0, 0)  # Preto
    p.setFont("Helvetica", 7)
    y -= 0.7*cm
    
    page_count = 1
    for parcela in parcelas:
        if y < 2*cm:
            # Rodapé
            p.setFont("Helvetica", 8)
            p.drawRightString(width - 1*cm, 1*cm, f"Página {page_count}")
            
            p.showPage()
            page_count += 1
            
            # Redesenhar cabeçalho na nova página
            draw_header(p, width, height)
            
            # Redesenhar cabeçalhos da tabela
            y = height - 4*cm
            p.setFillColorRGB(0.059, 0.090, 0.165)
            p.rect(0.5*cm, y - 0.3*cm, width - 1*cm, 0.6*cm, fill=True, stroke=False)
            
            p.setFillColorRGB(1, 1, 1)
            p.setFont("Helvetica-Bold", 8)
            p.drawString(0.7*cm, y, "Código")
            p.drawString(3*cm, y, "NF")
            p.drawString(5*cm, y, "Cliente")
            p.drawString(10*cm, y, "CNPJ/CPF")
            p.drawString(14.5*cm, y, "Tipo")
            p.drawString(16.5*cm, y, "Parc")
            p.drawString(18*cm, y, "Valor")
            p.drawString(21*cm, y, "Vencimento")
            p.drawString(24.5*cm, y, "Instituição")
            
            p.setFillColorRGB(0, 0, 0)
            p.setFont("Helvetica", 7)
            y -= 0.7*cm
        
        p.drawString(0.7*cm, y, str(parcela.codigo_identificador)[:12])
        p.drawString(3*cm, y, str(parcela.nota_fiscal.numero_nota))
        p.drawString(5*cm, y, parcela.cliente.nome[:22])
        p.drawString(10*cm, y, parcela.cliente.cpf_cnpj[:18])
        p.drawString(14.5*cm, y, parcela.tipo_parcela)
        p.drawString(16.5*cm, y, str(parcela.numero_parcela))
        p.drawString(18*cm, y, f"R$ {parcela.valor:,.2f}")
        p.drawString(21*cm, y, parcela.data_vencimento.strftime('%d/%m/%Y'))
        p.drawString(24.5*cm, y, (parcela.conta_financeira.nome[:12] if parcela.conta_financeira else '-'))
        
        y -= 0.5*cm
    
    # Rodapé da última página
    p.setFont("Helvetica", 8)
    p.drawRightString(width - 1*cm, 1*cm, f"Página {page_count}")
    
    p.showPage()
    p.save()
    
    return response


# ==================== NEGOCIAÇÃO DE PARCELAS ====================

@login_required
def processar_negociacao(request, parcela_id):
    """Processa a negociação de uma parcela com pagamento parcial"""
    import json
    from .models import HistoricoNegociacao, Parcela

    if request.method != 'POST':
        return JsonResponse({'error': 'Método não permitido'}, status=405)

    parcela = get_object_or_404(Parcela, id=parcela_id)

    try:
        # Receber dados JSON
        data = json.loads(request.body)

        valor_pago = Decimal(str(data.get('valor_pago', '0')))
        data_pagamento = datetime.strptime(data.get('data_pagamento'), '%Y-%m-%d').date()
        desconto = Decimal(str(data.get('desconto', '0')))
        juros = Decimal(str(data.get('juros', '0')))
        multa = Decimal(str(data.get('multa', '0')))
        juros_negociacao = Decimal(str(data.get('juros_negociacao', '0')))
        observacao = data.get('observacao', '')
        observacao_negociacao = data.get('observacao_negociacao', '')
        parcelas_dados = data.get('parcelas', [])
        is_consolidada = data.get('is_consolidada', False)
        parcelas_consolidadas = data.get('parcelas_consolidadas', [])
        # Garantir que parcelas_consolidadas seja lista de IDs
        if isinstance(parcelas_consolidadas, str):
            parcelas_consolidadas = [pid.strip() for pid in parcelas_consolidadas.split(',') if pid.strip()]
        elif not isinstance(parcelas_consolidadas, list):
            parcelas_consolidadas = []

        # Validação dos dados recebidos
        if not isinstance(parcelas_dados, list) or not parcelas_dados:
            return JsonResponse({'error': 'Nenhuma parcela informada ou formato inválido.'}, status=400)
        if is_consolidada and (not isinstance(parcelas_consolidadas, list) or not parcelas_consolidadas):
            return JsonResponse({'error': 'Negociação consolidada requer seleção de parcelas.'}, status=400)
        # Validar campos obrigatórios de cada parcela
        for parcela_info in parcelas_dados:
            if 'valor' not in parcela_info or 'vencimento' not in parcela_info or 'conta_financeira_id' not in parcela_info:
                return JsonResponse({'error': 'Campos obrigatórios faltando em parcelas.'}, status=400)

        sufixos = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        parcelas_criadas = []

        if is_consolidada and parcelas_consolidadas:
            # Negociação consolidada: processar todas as parcelas pendentes
            from .models import Parcela
            total_pago = valor_pago
            total_original = Decimal('0.00')
            ids_negociadas = []
            for pid in parcelas_consolidadas:
                p = Parcela.objects.get(id=pid)
                saldo_pendente = p.valor - (p.valor_pago or Decimal('0.00'))
                total_original += saldo_pendente
                # Atualizar parcela para -N
                p.codigo_identificador = p.codigo_identificador + '-N' if not p.codigo_identificador.endswith('-N') else p.codigo_identificador
                p.valor_pago = p.valor_pago or Decimal('0.00')
                # NÃO altere p.valor! Mantenha o valor original para referência no PDF
                p.data_pagamento = data_pagamento
                p.status_pagamento = 'pago'
                p.observacao = f"NEGOCIADA CONSOLIDADA: {observacao}"
                p.save()
                ids_negociadas.append(str(p.id))
                # Movimentação financeira
                if p.conta_financeira:
                    from financeiro.models import MovimentacaoFinanceira
                    descricao = f"Recebimento Parcial - {p.codigo_identificador}"
                    if p.cliente:
                        descricao += f" - {p.cliente.nome}"
                    MovimentacaoFinanceira.objects.create(
                        conta_financeira=p.conta_financeira,
                        data=data_pagamento,
                        tipo='ENTRADA',
                        valor=p.valor_pago,
                        descricao=descricao,
                        origem='CONTA_RECEBER',
                        observacoes=f"Negociação consolidada: {observacao_negociacao}",
                        usuario=request.user
                    )

            # Criar histórico de negociação consolidada
            from .models import HistoricoNegociacao
            historico = HistoricoNegociacao.objects.create(
                parcela_negociada=parcela,
                valor_original=total_original,
                valor_pago=valor_pago,
                saldo_renegociado=total_original - valor_pago + juros_negociacao,
                juros_negociacao=juros_negociacao,
                quantidade_parcelas=len(parcelas_dados),
                observacao=observacao_negociacao,
                usuario=request.user,
                is_consolidada=True,
                parcelas_consolidadas=','.join(ids_negociadas)
            )

            # Criar novas parcelas (A, B, C...)
            for i, parcela_info in enumerate(parcelas_dados):
                valor_nova = Decimal(str(parcela_info.get('valor', '0')))
                vencimento = datetime.strptime(parcela_info.get('vencimento'), '%Y-%m-%d').date()
                conta_financeira_id = parcela_info.get('conta_financeira_id')
                nova_parcela = Parcela.objects.create(
                    nota_fiscal=parcela.nota_fiscal,
                    cliente=parcela.cliente,
                    origem=parcela.origem,
                    conta_financeira_id=conta_financeira_id if conta_financeira_id else parcela.conta_financeira_id,
                    tipo_parcela=parcela.tipo_parcela,
                    numero_parcela=parcela.numero_parcela,
                    codigo_identificador=f"CONSOL-{parcela.cliente.id}-{sufixos[i]}",
                    valor=valor_nova,
                    valor_pago=0,
                    data_vencimento=vencimento,
                    status_pagamento='pendente',
                    observacao=f"Parcela gerada pela negociação consolidada de {','.join(ids_negociadas)}"
                )
                parcelas_criadas.append({
                    'codigo': nova_parcela.codigo_identificador,
                    'valor': str(nova_parcela.valor),
                    'vencimento': nova_parcela.data_vencimento.strftime('%d/%m/%Y')
                })

            return JsonResponse({
                'success': True,
                'message': f'Negociação consolidada registrada! Criadas {len(parcelas_criadas)} novas parcelas.',
                'parcelas_criadas': parcelas_criadas
            })

        else:
            # Negociação simples (original)
            saldo_restante = parcela.valor - valor_pago + juros_negociacao
            if saldo_restante <= 0:
                return JsonResponse({'error': 'Valor pago é maior ou igual ao valor da parcela'}, status=400)
            if not parcelas_dados:
                return JsonResponse({'error': 'Nenhuma parcela informada'}, status=400)
            valor_original = parcela.valor
            codigo_original = parcela.codigo_identificador
            parcela.codigo_identificador = codigo_original + '-N' if not codigo_original.endswith('-N') else codigo_original
            parcela.valor = valor_pago
            parcela.valor_pago = valor_pago
            parcela.data_pagamento = data_pagamento
            parcela.status_pagamento = 'pago'
            parcela.observacao = f"NEGOCIADA: {observacao}"
            parcela.save()
            if parcela.conta_financeira:
                from financeiro.models import MovimentacaoFinanceira
                descricao = f"Recebimento Parcial - {codigo_original}"
                if parcela.cliente:
                    descricao += f" - {parcela.cliente.nome}"
                MovimentacaoFinanceira.objects.create(
                    conta_financeira=parcela.conta_financeira,
                    data=data_pagamento,
                    tipo='ENTRADA',
                    valor=valor_pago,
                    descricao=descricao,
                    origem='CONTA_RECEBER',
                    observacoes=f"Negociação: {observacao_negociacao}",
                    usuario=request.user
                )
            historico = HistoricoNegociacao.objects.create(
                parcela_negociada=parcela,
                valor_original=valor_original,
                valor_pago=valor_pago,
                saldo_renegociado=saldo_restante,
                juros_negociacao=juros_negociacao,
                quantidade_parcelas=len(parcelas_dados),
                observacao=observacao_negociacao,
                usuario=request.user
            )
            for i, parcela_info in enumerate(parcelas_dados):
                valor_nova = Decimal(str(parcela_info.get('valor', '0')))
                vencimento = datetime.strptime(parcela_info.get('vencimento'), '%Y-%m-%d').date()
                conta_financeira_id = parcela_info.get('conta_financeira_id')
                nova_parcela = Parcela.objects.create(
                    nota_fiscal=parcela.nota_fiscal,
                    cliente=parcela.cliente,
                    origem=parcela.origem,
                    conta_financeira_id=conta_financeira_id if conta_financeira_id else parcela.conta_financeira_id,
                    tipo_parcela=parcela.tipo_parcela,
                    numero_parcela=parcela.numero_parcela,
                    codigo_identificador=f"{codigo_original}-{sufixos[i]}",
                    valor=valor_nova,
                    valor_pago=0,
                    data_vencimento=vencimento,
                    status_pagamento='pendente',
                    observacao=f"Parcela gerada pela negociação de {codigo_original}"
                )
                parcelas_criadas.append({
                    'codigo': nova_parcela.codigo_identificador,
                    'valor': str(nova_parcela.valor),
                    'vencimento': nova_parcela.data_vencimento.strftime('%d/%m/%Y')
                })
            return JsonResponse({
                'success': True,
                'message': f'Negociação registrada com sucesso! Criadas {len(parcelas_criadas)} novas parcelas.',
                'parcelas_criadas': parcelas_criadas
            })

    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Erro ao processar negociação: {str(e)}'}, status=500)

@login_required
def lista_historico_negociacoes(request):
    """Lista histórico de negociações de parcelas"""
    from .models import HistoricoNegociacao
    
    historicos = HistoricoNegociacao.objects.select_related(
        'parcela_negociada',
        'parcela_negociada__cliente',
        'parcela_negociada__nota_fiscal',
        'usuario'
    ).all()
    
    # Filtros
    cliente_id = request.GET.get('cliente', '').strip()
    usuario_id = request.GET.get('usuario', '').strip()
    data_inicio = request.GET.get('data_inicio', '').strip()
    data_fim = request.GET.get('data_fim', '').strip()
    
    if cliente_id:
        historicos = historicos.filter(parcela_negociada__cliente_id=cliente_id)
    if usuario_id:
        historicos = historicos.filter(usuario_id=usuario_id)
    if data_inicio:
        historicos = historicos.filter(data_negociacao__gte=data_inicio)
    if data_fim:
        historicos = historicos.filter(data_negociacao__lte=data_fim)
    
    # Resumo
    total_negociacoes = historicos.count()
    total_renegociado = historicos.aggregate(Sum('saldo_renegociado'))['saldo_renegociado__sum'] or Decimal('0.00')
    
    # Paginação
    paginator = Paginator(historicos, 20)
    page = request.GET.get('page')
    
    try:
        historicos_page = paginator.page(page)
    except PageNotAnInteger:
        historicos_page = paginator.page(1)
    except EmptyPage:
        historicos_page = paginator.page(paginator.num_pages)
    
    # Buscar usuários que fizeram negociações
    from django.contrib.auth.models import User
    usuarios = User.objects.filter(id__in=HistoricoNegociacao.objects.values_list('usuario_id', flat=True).distinct())
    
    context = {
        'hide_sidebar': True,
        'historicos': historicos_page,
        'page_obj': historicos_page,
        'clientes': Cliente.objects.filter(ativo=True),
        'usuarios': usuarios,
        'total_negociacoes': total_negociacoes,
        'total_renegociado': total_renegociado,
        'filtros': {
            'cliente': cliente_id,
            'usuario': usuario_id,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
        }
    }
    return render(request, 'contas_receber/lista_historico_negociacoes.html', context)


@login_required
def gerar_pdf_negociacoes(request):
    """Gera PDF do histórico de negociações"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from io import BytesIO
    from .models import HistoricoNegociacao
    
    # Filtros (mesmos da lista)
    historicos = HistoricoNegociacao.objects.select_related(
        'parcela_negociada',
        'parcela_negociada__cliente',
        'usuario'
    ).all()
    
    cliente_id = request.GET.get('cliente', '').strip()
    usuario_id = request.GET.get('usuario', '').strip()
    data_inicio = request.GET.get('data_inicio', '').strip()
    data_fim = request.GET.get('data_fim', '').strip()
    
    if cliente_id:
        historicos = historicos.filter(parcela_negociada__cliente_id=cliente_id)
    if usuario_id:
        historicos = historicos.filter(usuario_id=usuario_id)
    if data_inicio:
        historicos = historicos.filter(data_negociacao__gte=data_inicio)
    if data_fim:
        historicos = historicos.filter(data_negociacao__lte=data_fim)
    
    # Criar PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=2*cm, bottomMargin=1*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#0ea5e9'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    elements.append(Paragraph("RELATÓRIO DE NEGOCIAÇÕES DE PARCELAS", title_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Informações do filtro
    info_style = ParagraphStyle('Info', parent=styles['Normal'], fontSize=9, textColor=colors.grey, alignment=TA_CENTER)
    
    filtros_text = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    if data_inicio or data_fim:
        periodo = []
        if data_inicio:
            periodo.append(f"De: {datetime.strptime(data_inicio, '%Y-%m-%d').strftime('%d/%m/%Y')}")
        if data_fim:
            periodo.append(f"Até: {datetime.strptime(data_fim, '%Y-%m-%d').strftime('%d/%m/%Y')}")
        filtros_text += f" | Período: {' '.join(periodo)}"
    
    elements.append(Paragraph(filtros_text, info_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Resumo
    total_negociacoes = historicos.count()
    total_renegociado = historicos.aggregate(Sum('saldo_renegociado'))['saldo_renegociado__sum'] or Decimal('0.00')
    total_juros = historicos.aggregate(Sum('juros_negociacao'))['juros_negociacao__sum'] or Decimal('0.00')
    
    resumo_data = [
        ['Total de Negociações', 'Total Renegociado', 'Total de Juros'],
        [str(total_negociacoes), f'R$ {total_renegociado:,.2f}', f'R$ {total_juros:,.2f}']
    ]
    
    resumo_table = Table(resumo_data, colWidths=[6*cm, 6*cm, 6*cm])
    resumo_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0ea5e9')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    
    elements.append(resumo_table)
    elements.append(Spacer(1, 0.8*cm))
    
    # Tabela de dados
    data = [['Data', 'Parcela', 'Cliente', 'Valor Original', 'Valor Pago', 'Saldo', 'Juros', 'Qtd', 'Usuário']]
    
    for h in historicos[:100]:  # Limitar a 100 registros
        data.append([
            h.data_negociacao.strftime('%d/%m/%Y'),
            h.parcela_negociada.codigo_identificador or '-',
            h.parcela_negociada.cliente.nome[:20] if h.parcela_negociada.cliente else '-',
            f'R$ {h.valor_original:,.2f}',
            f'R$ {h.valor_pago:,.2f}',
            f'R$ {h.saldo_renegociado:,.2f}',
            f'R$ {h.juros_negociacao:,.2f}',
            str(h.quantidade_parcelas),
            h.usuario.username[:10] if h.usuario else '-'
        ])
    
    # Criar tabela
    col_widths = [2.5*cm, 2.5*cm, 4*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2*cm, 1.5*cm, 2.5*cm]
    table = Table(data, colWidths=col_widths)
    
    # Estilo da tabela
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # Rodapé
    elements.append(Spacer(1, 0.5*cm))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7, textColor=colors.grey, alignment=TA_CENTER)
    elements.append(Paragraph(f"Sistema de Gestão - Relatório gerado automaticamente", footer_style))
    
    # Gerar PDF
    doc.build(elements)
    
    # Retornar response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="historico_negociacoes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    return response


@login_required
def gerar_pdf_negociacao_individual(request, negociacao_id):
    """Gera PDF formal de uma negociação específica para envio ao cliente"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
    from io import BytesIO
    from .models import HistoricoNegociacao
    
    negociacao = get_object_or_404(HistoricoNegociacao, id=negociacao_id)
    parcela_original = negociacao.parcela_negociada
    cliente = parcela_original.cliente
    
    # Buscar parcelas geradas
    if negociacao.is_consolidada and negociacao.parcelas_consolidadas:
        # Buscar todas as parcelas originais consolidadas
        if negociacao.parcelas_consolidadas:
            ids_consolidadas = [int(pid) for pid in negociacao.parcelas_consolidadas.split(',') if pid]
            parcelas_consolidadas = Parcela.objects.filter(id__in=ids_consolidadas)
        else:
            parcelas_consolidadas = None
        # Buscar todas as novas parcelas geradas pela negociação consolidadas
        parcelas_geradas = Parcela.objects.filter(
            cliente=cliente,
            codigo_identificador__startswith=f'CONSOL-{cliente.id}-',
        ).order_by('data_vencimento')
    else:
        codigo_base = parcela_original.codigo_identificador.replace('-N', '')
        parcelas_consolidadas = None
        parcelas_geradas = Parcela.objects.filter(
            codigo_identificador__startswith=codigo_base,
            codigo_identificador__regex=r'-[A-Z]$'
        ).order_by('data_vencimento')
    
    # Criar PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2.5*cm, bottomMargin=2*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Cabeçalho - Título do Documento
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#0f172a'),
        spaceAfter=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#475569'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica'
    )
    
    elements.append(Paragraph("TERMO DE ACORDO DE PAGAMENTO", title_style))
    elements.append(Paragraph(f"Nº {negociacao.id:06d}/{datetime.now().year}", subtitle_style))
    elements.append(Spacer(1, 1*cm))
    
    # Estilo para texto do contrato
    body_style = ParagraphStyle(
        'Body',
        parent=styles['Normal'],
        fontSize=11,
        leading=16,
        alignment=TA_JUSTIFY,
        fontName='Helvetica'
    )
    
    bold_style = ParagraphStyle(
        'Bold',
        parent=styles['Normal'],
        fontSize=11,
        fontName='Helvetica-Bold'
    )
    
    # Partes do Acordo
    elements.append(Paragraph("<b>DAS PARTES:</b>", bold_style))
    elements.append(Spacer(1, 0.3*cm))
    
    partes_text = f"""
    <b>CREDOR:</b> Folia Imports Ltda.<br/>
    <br/>
    <b>DEVEDOR:</b> {cliente.nome}<br/>
    <b>CPF/CNPJ:</b> {cliente.cpf_cnpj}<br/>
    """
    
    elements.append(Paragraph(partes_text, body_style))
    elements.append(Spacer(1, 0.8*cm))
    
    # Do Objeto
    elements.append(Paragraph("<b>DO OBJETO:</b>", bold_style))
    elements.append(Spacer(1, 0.3*cm))
    if negociacao.is_consolidada:
        if parcelas_consolidadas:
            # Montar tabela de duas colunas: Código e Valor Original
            objeto_text = "Pelo presente instrumento, as partes acima qualificadas acordam a renegociação consolidada de débito(s) em aberto, referente às seguintes parcelas:"
            elements.append(Paragraph(objeto_text, body_style))
            elements.append(Spacer(1, 0.3*cm))
            # Dados da tabela
            table_data = [['Parcela', 'Vencimento', 'Valor Original']]
            for p in parcelas_consolidadas:
                table_data.append([
                    f"{p.codigo_identificador}",
                    p.data_vencimento.strftime('%d/%m/%Y'),
                    f"R$ {p.valor:,.2f}"
                ])
            table = Table(table_data, colWidths=[5*cm, 4*cm, 4*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 0.5*cm))
            elements.append(Paragraph("nas condições estabelecidas neste termo.", body_style))
            elements.append(Spacer(1, 0.8*cm))
            objeto_text = None
        else:
            # Caso não haja parcelas consolidadas, mostrar mensagem alternativa
            objeto_text = "Pelo presente instrumento, as partes acima qualificadas acordam a renegociação consolidada de débito(s) em aberto. (Nenhuma parcela consolidada encontrada para exibir.)"
            elements.append(Paragraph(objeto_text, body_style))
            elements.append(Spacer(1, 0.8*cm))
    else:
        objeto_text = f"""
        Pelo presente instrumento, as partes acima qualificadas acordam a renegociação de débito 
        em aberto, referente à parcela <b>{codigo_base}</b>, nas condições estabelecidas neste termo.
        """
    if objeto_text:
        elements.append(Paragraph(objeto_text, body_style))
        elements.append(Spacer(1, 0.8*cm))
    
    # Dos Valores
    elements.append(Paragraph("<b>DOS VALORES:</b>", bold_style))
    elements.append(Spacer(1, 0.3*cm))
    
    valores_data = [
        ['Descrição', 'Valor'],
        ['Valor Original do Débito', f'R$ {negociacao.valor_original:,.2f}'],
        ['Valor Pago em {}'.format(negociacao.data_negociacao.strftime('%d/%m/%Y')), f'R$ {negociacao.valor_pago:,.2f}'],
    ]
    
    if negociacao.juros_negociacao > 0:
        valores_data.append(['Juros Aplicados', f'R$ {negociacao.juros_negociacao:,.2f}'])
    
    valores_data.append(['Saldo Renegociado', f'R$ {negociacao.saldo_renegociado:,.2f}'])
    
    valores_table = Table(valores_data, colWidths=[10*cm, 5*cm])
    valores_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('LINEBELOW', (0, -1), (-1, -1), 2, colors.HexColor('#0f172a')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    
    elements.append(valores_table)
    elements.append(Spacer(1, 0.8*cm))
    
    # Do Parcelamento
    elements.append(Paragraph("<b>DO PARCELAMENTO:</b>", bold_style))
    elements.append(Spacer(1, 0.3*cm))
    
    parcelamento_text = f"""
    O saldo devedor será quitado em <b>{negociacao.quantidade_parcelas} ({negociacao.quantidade_parcelas}) parcela(s)</b>, 
    conforme discriminado abaixo:
    """
    
    elements.append(Paragraph(parcelamento_text, body_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Tabela de Parcelas
    parcelas_data = [['Parcela', 'Vencimento', 'Valor', 'Instituição Financeira']]
    
    for i, parcela in enumerate(parcelas_geradas, 1):
        parcelas_data.append([
            f'{i}ª Parcela',
            parcela.data_vencimento.strftime('%d/%m/%Y'),
            f'R$ {parcela.valor:,.2f}',
            parcela.conta_financeira.nome if parcela.conta_financeira else '-'
        ])
    
    parcelas_table = Table(parcelas_data, colWidths=[3*cm, 3.5*cm, 3.5*cm, 5*cm])
    parcelas_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(parcelas_table)
    elements.append(Spacer(1, 0.8*cm))
    
    # Das Condições
    elements.append(Paragraph("<b>DAS CONDIÇÕES:</b>", bold_style))
    elements.append(Spacer(1, 0.3*cm))
    
    condicoes_text = """
    1. O não pagamento de qualquer parcela na data de vencimento implicará no vencimento 
    antecipado de todas as demais parcelas, ficando o débito total sujeito à cobrança judicial.<br/>
    <br/>
    2. O devedor declara estar ciente de todas as condições estabelecidas neste acordo e 
    compromete-se a honrá-las integralmente.<br/>
    <br/>
    3. Este acordo é feito em caráter irrevogável e irretratável, constituindo título 
    executivo extrajudicial.
    """
    
    elements.append(Paragraph(condicoes_text, body_style))
    elements.append(Spacer(1, 1.5*cm))
    
    # Observações
    if negociacao.observacao:
        elements.append(Paragraph("<b>OBSERVAÇÕES:</b>", bold_style))
        elements.append(Spacer(1, 0.3*cm))
        elements.append(Paragraph(negociacao.observacao, body_style))
        elements.append(Spacer(1, 1*cm))
    
    # Data e Local
    data_local = f"""
    <br/><br/>
    ____________________, _____ de __________________ de {datetime.now().year}
    """
    elements.append(Paragraph(data_local, body_style))
    elements.append(Spacer(1, 2*cm))
    
    # Assinaturas
    assinatura_style = ParagraphStyle(
        'Assinatura',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        fontName='Helvetica'
    )
    
    assinaturas_data = [
        ['_______________________________', '_______________________________'],
        ['Credor', 'Devedor'],
        ['', f'{cliente.nome}'],
        ['', f'CPF/CNPJ: {cliente.cpf_cnpj}']
    ]
    
    assinaturas_table = Table(assinaturas_data, colWidths=[7.5*cm, 7.5*cm])
    assinaturas_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 0),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
    ]))
    
    elements.append(assinaturas_table)
    
    # Rodapé
    elements.append(Spacer(1, 1*cm))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
    elements.append(Paragraph(f"Documento gerado eletronicamente em {datetime.now().strftime('%d/%m/%Y às %H:%M')}", footer_style))
    
    # Gerar PDF
    doc.build(elements)
    
    # Retornar response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="acordo_pagamento_{negociacao.id:06d}_{cliente.cpf_cnpj}.pdf"'
    
    return response


@login_required
def verificar_parcelas_cliente(request, cliente_id):
    """Verifica se o cliente tem outras parcelas pendentes"""
    parcelas_pendentes = Parcela.objects.filter(
        cliente_id=cliente_id,
        status_pagamento__in=['pendente', 'parcial']
    ).exclude(
        id=request.GET.get('excluir_id')  # Excluir a parcela atual
    ).values('id', 'codigo_identificador', 'valor', 'valor_pago', 'data_vencimento')
    
    outras_parcelas = []
    total_outras = Decimal('0.00')
    
    for p in parcelas_pendentes:
        saldo = Decimal(str(p['valor'])) - Decimal(str(p['valor_pago'] or 0))
        outras_parcelas.append({
            'id': p['id'],
            'codigo': p['codigo_identificador'],
            'saldo': float(saldo),
            'vencimento': p['data_vencimento'].strftime('%d/%m/%Y')
        })
        total_outras += saldo
    
    return JsonResponse({
        'outras_parcelas': outras_parcelas,
        'total_outras': float(total_outras),
        'quantidade': len(outras_parcelas)
    })


# ==================== RELATÓRIOS PDF ====================

@login_required
def menu_relatorios(request):
    """Menu principal de relatórios"""
    from datetime import date
    
    clientes = Cliente.objects.all().order_by('nome')
    
    context = {
        'hide_sidebar': True,
        'clientes': clientes,
        'hoje': date.today(),
    }
    return render(request, 'contas_receber/menu_relatorios.html', context)


@login_required
def relatorio_titulos_vencer_pdf(request):
    """Gera PDF de títulos a vencer"""
    from .relatorios_pdf import RelatorioTitulosVencer
    
    dias = int(request.GET.get('dias', 30))
    cliente_id = request.GET.get('cliente')
    
    relatorio = RelatorioTitulosVencer()
    buffer = relatorio.gerar(dias=dias, cliente_id=cliente_id)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="titulos_a_vencer_{date.today()}.pdf"'
    return response


@login_required
def relatorio_titulos_vencidos_pdf(request):
    """Gera PDF de títulos vencidos (inadimplência)"""
    from .relatorios_pdf import RelatorioTitulosVencidos
    
    cliente_id = request.GET.get('cliente')
    
    relatorio = RelatorioTitulosVencidos()
    buffer = relatorio.gerar(cliente_id=cliente_id)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="inadimplencia_{date.today()}.pdf"'
    return response


@login_required
def relatorio_recebimentos_pdf(request):
    """Gera PDF de recebimentos realizados"""
    from .relatorios_pdf import RelatorioRecebimentos
    
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    cliente_id = request.GET.get('cliente')
    
    if not data_inicio or not data_fim:
        # Padrão: mês atual
        data_inicio = date.today().replace(day=1)
        ultimo_dia = (data_inicio + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        data_fim = ultimo_dia
    else:
        data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
    
    relatorio = RelatorioRecebimentos()
    buffer = relatorio.gerar(data_inicio, data_fim, cliente_id)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="recebimentos_{data_inicio}_{data_fim}.pdf"'
    return response


@login_required
def relatorio_fluxo_caixa_pdf(request):
    """Gera PDF de fluxo de caixa projetado"""
    from .relatorios_pdf import RelatorioFluxoCaixa
    
    dias = int(request.GET.get('dias', 60))
    
    relatorio = RelatorioFluxoCaixa()
    buffer = relatorio.gerar(dias=dias)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="fluxo_caixa_projetado_{date.today()}.pdf"'
    return response


@login_required
def relatorio_por_cliente_pdf(request, cliente_id):
    """Gera PDF completo de um cliente"""
    from .relatorios_pdf import RelatorioPorCliente
    
    relatorio = RelatorioPorCliente()
    buffer = relatorio.gerar(cliente_id)
    
    cliente = Cliente.objects.get(id=cliente_id)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="cliente_{cliente.cpf_cnpj}_{date.today()}.pdf"'
    return response


# ==================== EXPORTAÇÃO EXCEL ====================

@login_required
def relatorio_titulos_vencer_excel(request):
    """Exporta relatório de títulos a vencer para Excel"""
    from .relatorios_excel import RelatorioTitulosVencerExcel
    from io import BytesIO
    
    dias = int(request.GET.get('dias', 30))
    cliente_id = request.GET.get('cliente')
    
    # Buscar parcelas
    parcelas = Parcela.objects.filter(
        status_pagamento='pendente',
        data_vencimento__gte=date.today(),
        data_vencimento__lte=date.today() + timedelta(days=dias)
    ).select_related('nota_fiscal', 'nota_fiscal__cliente').order_by('data_vencimento')
    
    if cliente_id:
        parcelas = parcelas.filter(nota_fiscal__cliente_id=cliente_id)
    
    # Gerar Excel
    usuario = f"{request.user.first_name} {request.user.last_name}" if request.user.first_name else request.user.username
    relatorio = RelatorioTitulosVencerExcel(parcelas, dias, usuario)
    wb = relatorio.gerar()
    
    # Salvar em buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Retornar response
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="titulos_a_vencer_{date.today()}.xlsx"'
    return response


@login_required
def relatorio_titulos_vencidos_excel(request):
    """Exporta relatório de inadimplência para Excel"""
    from .relatorios_excel import RelatorioTitulosVencidosExcel
    from io import BytesIO
    
    cliente_id = request.GET.get('cliente')
    
    # Buscar parcelas vencidas
    parcelas = Parcela.objects.filter(
        status_pagamento='pendente',
        data_vencimento__lt=date.today()
    ).select_related('nota_fiscal', 'nota_fiscal__cliente').order_by('data_vencimento')
    
    if cliente_id:
        parcelas = parcelas.filter(nota_fiscal__cliente_id=cliente_id)
    
    # Gerar Excel
    usuario = f"{request.user.first_name} {request.user.last_name}" if request.user.first_name else request.user.username
    relatorio = RelatorioTitulosVencidosExcel(parcelas, usuario)
    wb = relatorio.gerar()
    
    # Salvar em buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Retornar response
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="inadimplencia_{date.today()}.xlsx"'
    return response


@login_required
def relatorio_recebimentos_excel(request):
    """Exporta relatório de recebimentos para Excel"""
    from .relatorios_excel import RelatorioRecebimentosExcel
    from io import BytesIO
    
    # Período padrão: mês atual
    hoje = date.today()
    data_inicio = request.GET.get('data_inicio', f"{hoje.year}-{hoje.month:02d}-01")
    data_fim = request.GET.get('data_fim', hoje.strftime('%Y-%m-%d'))
    
    if isinstance(data_inicio, str):
        data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
    if isinstance(data_fim, str):
        data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
    
    # Buscar parcelas pagas no período
    parcelas = Parcela.objects.filter(
        status_pagamento='pago',
        data_pagamento__gte=data_inicio,
        data_pagamento__lte=data_fim
    ).select_related('nota_fiscal', 'nota_fiscal__cliente').order_by('data_pagamento')
    
    # Gerar Excel
    usuario = f"{request.user.first_name} {request.user.last_name}" if request.user.first_name else request.user.username
    relatorio = RelatorioRecebimentosExcel(parcelas, data_inicio, data_fim, usuario)
    wb = relatorio.gerar()
    
    # Salvar em buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Retornar response
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="recebimentos_{data_inicio}_{data_fim}.xlsx"'
    return response


@login_required
def relatorio_fluxo_caixa_excel(request):
    """Exporta relatório de fluxo de caixa para Excel"""
    from .relatorios_excel import RelatorioFluxoCaixaExcel
    from io import BytesIO
    
    dias = int(request.GET.get('dias', 60))
    
    # Buscar parcelas futuras
    parcelas = Parcela.objects.filter(
        status_pagamento='pendente',
        data_vencimento__gte=date.today(),
        data_vencimento__lte=date.today() + timedelta(days=dias)
    ).select_related('nota_fiscal', 'nota_fiscal__cliente').order_by('data_vencimento')
    
    # Gerar Excel
    usuario = f"{request.user.first_name} {request.user.last_name}" if request.user.first_name else request.user.username
    relatorio = RelatorioFluxoCaixaExcel(parcelas, dias, usuario)
    wb = relatorio.gerar()
    
    # Salvar em buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Retornar response
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="fluxo_caixa_{date.today()}.xlsx"'
    return response


@login_required
def relatorio_por_cliente_excel(request, cliente_id):
    """Exporta relatório de cliente para Excel"""
    from .relatorios_excel import RelatorioPorClienteExcel
    from io import BytesIO
    
    cliente = Cliente.objects.get(id=cliente_id)
    
    # Estatísticas
    parcelas_todas = Parcela.objects.filter(nota_fiscal__cliente=cliente)
    parcelas_pendentes = parcelas_todas.filter(status_pagamento='pendente', data_vencimento__gte=date.today())
    parcelas_vencidas = parcelas_todas.filter(status_pagamento='pendente', data_vencimento__lt=date.today())
    parcelas_pagas = parcelas_todas.filter(status_pagamento='pago').order_by('-data_pagamento')
    
    estatisticas = {
        'total_parcelas': parcelas_todas.count(),
        'pendentes': parcelas_pendentes.count(),
        'vencidas': parcelas_vencidas.count(),
        'pagas': parcelas_pagas.count(),
        'total_pendente': parcelas_pendentes.aggregate(total=Sum('valor'))['total'] or Decimal('0'),
        'total_vencido': parcelas_vencidas.aggregate(total=Sum('valor'))['total'] or Decimal('0'),
        'total_pago': parcelas_pagas.aggregate(total=Sum('valor'))['total'] or Decimal('0'),
    }
    
    # Gerar Excel
    usuario = f"{request.user.first_name} {request.user.last_name}" if request.user.first_name else request.user.username
    relatorio = RelatorioPorClienteExcel(
        cliente, 
        estatisticas, 
        parcelas_pendentes.select_related('nota_fiscal'), 
        parcelas_vencidas.select_related('nota_fiscal'),
        parcelas_pagas.select_related('nota_fiscal'),
        usuario
    )
    wb = relatorio.gerar()
    
    # Salvar em buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Retornar response
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="cliente_{cliente.cpf_cnpj}_{date.today()}.xlsx"'
    return response


@login_required
def relatorio_extrato_periodo_pdf(request):
    """Gera relatório de Extrato por Período em PDF"""
    from .relatorios_pdf import RelatorioExtratoPeriodo
    from io import BytesIO
    
    # Obter parâmetros de data
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    
    # Converter para objetos date
    if data_inicio:
        data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
    else:
        data_inicio = date.today().replace(day=1)  # Primeiro dia do mês
    
    if data_fim:
        data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
    else:
        data_fim = date.today()
    
    # Gerar PDF
    relatorio = RelatorioExtratoPeriodo()
    buffer = relatorio.gerar(data_inicio, data_fim)
    
    # Retornar response
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="extrato_periodo_{data_inicio}_{data_fim}.pdf"'
    return response


@login_required
def relatorio_extrato_periodo_excel(request):
    """Exporta relatório de Extrato por Período para Excel"""
    from .relatorios_excel import RelatorioExtratoPeriodoExcel
    from io import BytesIO
    
    # Obter parâmetros de data
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    
    # Converter para objetos date
    if data_inicio:
        data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
    else:
        data_inicio = date.today().replace(day=1)  # Primeiro dia do mês
    
    if data_fim:
        data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
    else:
        data_fim = date.today()
    
    # Gerar Excel
    usuario = f"{request.user.first_name} {request.user.last_name}" if request.user.first_name else request.user.username
    relatorio = RelatorioExtratoPeriodoExcel(usuario)
    wb = relatorio.gerar(data_inicio, data_fim)
    
    # Salvar em buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Retornar response
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="extrato_periodo_{data_inicio}_{data_fim}.xlsx"'
    return response
