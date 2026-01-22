"""
Módulo para exportação de relatórios em Excel
Gera planilhas formatadas profissionalmente com openpyxl
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from decimal import Decimal


class RelatorioExcelBase:
    """Classe base para relatórios Excel"""
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.linha_atual = 1
        
        # Estilos
        self.estilo_titulo = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        self.estilo_subtitulo = Font(name='Calibri', size=11, italic=True, color='666666')
        self.estilo_cabecalho = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        self.estilo_dados = Font(name='Calibri', size=10)
        self.estilo_total = Font(name='Calibri', size=11, bold=True)
        
        self.fill_titulo = PatternFill(start_color='0369A1', end_color='0369A1', fill_type='solid')
        self.fill_cabecalho = PatternFill(start_color='0EA5E9', end_color='0EA5E9', fill_type='solid')
        self.fill_total = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid')
        self.fill_vencido = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
        self.fill_alerta = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
        
        self.border_thin = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        self.align_center = Alignment(horizontal='center', vertical='center')
        self.align_left = Alignment(horizontal='left', vertical='center')
        self.align_right = Alignment(horizontal='right', vertical='center')
    
    def formatar_moeda(self, valor):
        """Formata valor como moeda brasileira"""
        if valor is None:
            return "R$ 0,00"
        return f"R$ {float(valor):,.2f}".replace(',', '_').replace('.', ',').replace('_', '.')
    
    def formatar_data(self, data):
        """Formata data no padrão brasileiro"""
        if data is None:
            return ""
        if isinstance(data, str):
            return data
        return data.strftime('%d/%m/%Y')
    
    def adicionar_cabecalho(self, titulo, subtitulo, usuario):
        """Adiciona cabeçalho padrão ao relatório"""
        # Título
        self.ws.merge_cells(f'A{self.linha_atual}:F{self.linha_atual}')
        cell = self.ws[f'A{self.linha_atual}']
        cell.value = titulo
        cell.font = self.estilo_titulo
        cell.fill = self.fill_titulo
        cell.alignment = self.align_center
        self.linha_atual += 1
        
        # Subtítulo
        self.ws.merge_cells(f'A{self.linha_atual}:F{self.linha_atual}')
        cell = self.ws[f'A{self.linha_atual}']
        cell.value = subtitulo
        cell.font = self.estilo_subtitulo
        cell.alignment = self.align_center
        self.linha_atual += 1
        
        # Info de emissão
        self.ws.merge_cells(f'A{self.linha_atual}:F{self.linha_atual}')
        cell = self.ws[f'A{self.linha_atual}']
        data_emissao = datetime.now().strftime('%d/%m/%Y às %H:%M')
        cell.value = f"Emitido em: {data_emissao} | Por: {usuario}"
        cell.font = Font(name='Calibri', size=9, italic=True, color='666666')
        cell.alignment = self.align_center
        self.linha_atual += 2  # Pula linha
    
    def adicionar_linha_cabecalho(self, colunas):
        """Adiciona linha de cabeçalho da tabela"""
        for idx, coluna in enumerate(colunas, start=1):
            cell = self.ws.cell(row=self.linha_atual, column=idx)
            cell.value = coluna
            cell.font = self.estilo_cabecalho
            cell.fill = self.fill_cabecalho
            cell.alignment = self.align_center
            cell.border = self.border_thin
        self.linha_atual += 1
    
    def adicionar_linha_dados(self, valores, destacar=None):
        """Adiciona linha de dados"""
        for idx, valor in enumerate(valores, start=1):
            cell = self.ws.cell(row=self.linha_atual, column=idx)
            cell.value = valor
            cell.font = self.estilo_dados
            cell.border = self.border_thin
            
            # Aplicar destaque se necessário
            if destacar == 'vencido':
                cell.fill = self.fill_vencido
            elif destacar == 'alerta':
                cell.fill = self.fill_alerta
            
            # Alinhamento baseado no tipo
            if isinstance(valor, (int, float, Decimal)) or (isinstance(valor, str) and valor.startswith('R$')):
                cell.alignment = self.align_right
            else:
                cell.alignment = self.align_left
        
        self.linha_atual += 1
    
    def adicionar_linha_total(self, label, valor, coluna_inicio=1, coluna_valor=6):
        """Adiciona linha de total"""
        # Merge células para label
        if coluna_valor > coluna_inicio + 1:
            self.ws.merge_cells(
                start_row=self.linha_atual,
                start_column=coluna_inicio,
                end_row=self.linha_atual,
                end_column=coluna_valor - 1
            )
        
        cell_label = self.ws.cell(row=self.linha_atual, column=coluna_inicio)
        cell_label.value = label
        cell_label.font = self.estilo_total
        cell_label.fill = self.fill_total
        cell_label.alignment = self.align_right
        cell_label.border = self.border_thin
        
        cell_valor = self.ws.cell(row=self.linha_atual, column=coluna_valor)
        cell_valor.value = valor
        cell_valor.font = self.estilo_total
        cell_valor.fill = self.fill_total
        cell_valor.alignment = self.align_right
        cell_valor.border = self.border_thin
        
        self.linha_atual += 1
    
    def ajustar_largura_colunas(self, larguras=None):
        """Ajusta largura das colunas"""
        if larguras:
            for idx, largura in enumerate(larguras, start=1):
                self.ws.column_dimensions[get_column_letter(idx)].width = largura
        else:
            # Auto-ajuste baseado no conteúdo
            for column in self.ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                self.ws.column_dimensions[column_letter].width = adjusted_width
    
    def gerar(self):
        """Retorna o workbook gerado"""
        return self.wb


class RelatorioTitulosVencerExcel(RelatorioExcelBase):
    """Relatório de títulos a vencer em Excel"""
    
    def __init__(self, parcelas, dias, usuario):
        super().__init__()
        self.ws.title = "Títulos a Vencer"
        
        # Cabeçalho
        titulo = "RELATÓRIO - TÍTULOS A VENCER"
        subtitulo = f"Próximos {dias} dias"
        self.adicionar_cabecalho(titulo, subtitulo, usuario)
        
        # Cabeçalho da tabela
        colunas = ['NF', 'Cliente', 'Parcela', 'Vencimento', 'Dias', 'Valor']
        self.adicionar_linha_cabecalho(colunas)
        
        # Dados
        total = Decimal('0')
        for parcela in parcelas:
            dias_vencer = (parcela.data_vencimento - date.today()).days
            
            valores = [
                parcela.nota_fiscal.numero_nota if parcela.nota_fiscal else '-',
                parcela.nota_fiscal.cliente.nome if parcela.nota_fiscal and parcela.nota_fiscal.cliente else '-',
                parcela.numero_parcela,
                self.formatar_data(parcela.data_vencimento),
                dias_vencer,
                self.formatar_moeda(parcela.valor)
            ]
            
            # Destacar se vence em menos de 7 dias
            destacar = 'alerta' if dias_vencer <= 7 else None
            self.adicionar_linha_dados(valores, destacar)
            total += parcela.valor
        
        # Total
        self.adicionar_linha_total('TOTAL:', self.formatar_moeda(total))
        
        # Ajustar larguras
        self.ajustar_largura_colunas([15, 35, 12, 15, 10, 18])


class RelatorioTitulosVencidosExcel(RelatorioExcelBase):
    """Relatório de inadimplência em Excel"""
    
    def __init__(self, parcelas, usuario):
        super().__init__()
        self.ws.title = "Inadimplência"
        
        # Cabeçalho
        titulo = "RELATÓRIO DE INADIMPLÊNCIA"
        subtitulo = "Títulos Vencidos e Não Pagos"
        self.adicionar_cabecalho(titulo, subtitulo, usuario)
        
        # Cabeçalho da tabela
        colunas = ['NF', 'Cliente', 'Parcela', 'Vencimento', 'Dias Atraso', 'Valor']
        self.adicionar_linha_cabecalho(colunas)
        
        # Dados
        total = Decimal('0')
        for parcela in parcelas:
            dias_atraso = (date.today() - parcela.data_vencimento).days
            
            valores = [
                parcela.nota_fiscal.numero_nota if parcela.nota_fiscal else '-',
                parcela.nota_fiscal.cliente.nome if parcela.nota_fiscal and parcela.nota_fiscal.cliente else '-',
                parcela.numero_parcela,
                self.formatar_data(parcela.data_vencimento),
                dias_atraso,
                self.formatar_moeda(parcela.valor)
            ]
            
            self.adicionar_linha_dados(valores, destacar='vencido')
            total += parcela.valor
        
        # Total
        self.adicionar_linha_total('TOTAL EM ATRASO:', self.formatar_moeda(total))
        
        # Ajustar larguras
        self.ajustar_largura_colunas([15, 35, 12, 15, 15, 18])


class RelatorioRecebimentosExcel(RelatorioExcelBase):
    """Relatório de recebimentos em Excel"""
    
    def __init__(self, parcelas, data_inicio, data_fim, usuario):
        super().__init__()
        self.ws.title = "Recebimentos"
        
        # Cabeçalho
        titulo = "RELATÓRIO DE RECEBIMENTOS"
        subtitulo = f"Período: {self.formatar_data(data_inicio)} a {self.formatar_data(data_fim)}"
        self.adicionar_cabecalho(titulo, subtitulo, usuario)
        
        # Cabeçalho da tabela
        colunas = ['NF', 'Cliente', 'Parcela', 'Vencimento', 'Pagamento', 'Valor']
        self.adicionar_linha_cabecalho(colunas)
        
        # Dados
        total = Decimal('0')
        for parcela in parcelas:
            valores = [
                parcela.nota_fiscal.numero_nota if parcela.nota_fiscal else '-',
                parcela.nota_fiscal.cliente.nome if parcela.nota_fiscal and parcela.nota_fiscal.cliente else '-',
                parcela.numero_parcela,
                self.formatar_data(parcela.data_vencimento),
                self.formatar_data(parcela.data_pagamento),
                self.formatar_moeda(parcela.valor)
            ]
            
            self.adicionar_linha_dados(valores)
            total += parcela.valor
        
        # Total
        self.adicionar_linha_total('TOTAL RECEBIDO:', self.formatar_moeda(total))
        
        # Ajustar larguras
        self.ajustar_largura_colunas([15, 35, 12, 15, 15, 18])


class RelatorioExtratoPeriodoExcel(RelatorioExcelBase):
    """Relatório de extrato completo por período em Excel"""
    
    def __init__(self, data_inicio, data_fim, parcelas_pagas, parcelas_vencidas, parcelas_a_vencer, usuario):
        super().__init__()
        self.ws.title = "Extrato Período"
        
        # Cabeçalho
        titulo = "EXTRATO POR PERÍODO"
        subtitulo = f"Período: {self.formatar_data(data_inicio)} a {self.formatar_data(data_fim)}"
        self.adicionar_cabecalho(titulo, subtitulo, usuario)
        
        # ===== PARCELAS PAGAS =====
        if parcelas_pagas:
            self.ws.merge_cells(f'A{self.linha_atual}:F{self.linha_atual}')
            cell = self.ws[f'A{self.linha_atual}']
            cell.value = "PARCELAS PAGAS"
            cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
            cell.alignment = self.align_center
            self.linha_atual += 1
            
            colunas = ['NF', 'Cliente', 'Parcela', 'Vencimento', 'Pagamento', 'Valor']
            self.adicionar_linha_cabecalho(colunas)
            
            total_pago = Decimal('0')
            for parcela in parcelas_pagas:
                valores = [
                    parcela.nota_fiscal.numero_nota if parcela.nota_fiscal else '-',
                    parcela.nota_fiscal.cliente.nome if parcela.nota_fiscal and parcela.nota_fiscal.cliente else '-',
                    parcela.numero_parcela,
                    self.formatar_data(parcela.data_vencimento),
                    self.formatar_data(parcela.data_pagamento),
                    self.formatar_moeda(parcela.valor)
                ]
                self.adicionar_linha_dados(valores)
                total_pago += parcela.valor
            
            self.adicionar_linha_total('TOTAL PAGO:', self.formatar_moeda(total_pago))
            self.linha_atual += 2
        
        # ===== PARCELAS VENCIDAS =====
        if parcelas_vencidas:
            self.ws.merge_cells(f'A{self.linha_atual}:F{self.linha_atual}')
            cell = self.ws[f'A{self.linha_atual}']
            cell.value = "PARCELAS VENCIDAS (INADIMPLÊNCIA)"
            cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
            cell.alignment = self.align_center
            self.linha_atual += 1
            
            colunas = ['NF', 'Cliente', 'Parcela', 'Vencimento', 'Dias Atraso', 'Valor']
            self.adicionar_linha_cabecalho(colunas)
            
            total_vencido = Decimal('0')
            for parcela in parcelas_vencidas:
                dias_atraso = (date.today() - parcela.data_vencimento).days
                valores = [
                    parcela.nota_fiscal.numero_nota if parcela.nota_fiscal else '-',
                    parcela.nota_fiscal.cliente.nome if parcela.nota_fiscal and parcela.nota_fiscal.cliente else '-',
                    parcela.numero_parcela,
                    self.formatar_data(parcela.data_vencimento),
                    dias_atraso,
                    self.formatar_moeda(parcela.valor)
                ]
                self.adicionar_linha_dados(valores, destacar='vencido')
                total_vencido += parcela.valor
            
            self.adicionar_linha_total('TOTAL VENCIDO:', self.formatar_moeda(total_vencido))
            self.linha_atual += 2
        
        # ===== PARCELAS A VENCER =====
        if parcelas_a_vencer:
            self.ws.merge_cells(f'A{self.linha_atual}:F{self.linha_atual}')
            cell = self.ws[f'A{self.linha_atual}']
            cell.value = "PARCELAS A VENCER"
            cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
            cell.alignment = self.align_center
            self.linha_atual += 1
            
            colunas = ['NF', 'Cliente', 'Parcela', 'Vencimento', 'Dias', 'Valor']
            self.adicionar_linha_cabecalho(colunas)
            
            total_a_vencer = Decimal('0')
            for parcela in parcelas_a_vencer:
                dias_vencer = (parcela.data_vencimento - date.today()).days
                valores = [
                    parcela.nota_fiscal.numero_nota if parcela.nota_fiscal else '-',
                    parcela.nota_fiscal.cliente.nome if parcela.nota_fiscal and parcela.nota_fiscal.cliente else '-',
                    parcela.numero_parcela,
                    self.formatar_data(parcela.data_vencimento),
                    dias_vencer,
                    self.formatar_moeda(parcela.valor)
                ]
                destacar = 'alerta' if dias_vencer <= 7 else None
                self.adicionar_linha_dados(valores, destacar)
                total_a_vencer += parcela.valor
            
            self.adicionar_linha_total('TOTAL A VENCER:', self.formatar_moeda(total_a_vencer))
            self.linha_atual += 2
        
        # ===== RESUMO GERAL =====
        self.ws.merge_cells(f'A{self.linha_atual}:F{self.linha_atual}')
        cell = self.ws[f'A{self.linha_atual}']
        cell.value = "RESUMO DO PERÍODO"
        cell.font = Font(name='Calibri', size=12, bold=True, color='0369A1')
        cell.alignment = self.align_center
        self.linha_atual += 1
        
        total_geral_pago = sum(p.valor for p in parcelas_pagas) if parcelas_pagas else Decimal('0')
        total_geral_vencido = sum(p.valor for p in parcelas_vencidas) if parcelas_vencidas else Decimal('0')
        total_geral_a_vencer = sum(p.valor for p in parcelas_a_vencer) if parcelas_a_vencer else Decimal('0')
        
        resumo = [
            ['Qtd Pagas:', len(parcelas_pagas), 'Valor Pago:', self.formatar_moeda(total_geral_pago)],
            ['Qtd Vencidas:', len(parcelas_vencidas), 'Valor Vencido:', self.formatar_moeda(total_geral_vencido)],
            ['Qtd A Vencer:', len(parcelas_a_vencer), 'Valor A Vencer:', self.formatar_moeda(total_geral_a_vencer)],
        ]
        
        for linha_resumo in resumo:
            for idx, valor in enumerate(linha_resumo, start=1):
                cell = self.ws.cell(row=self.linha_atual, column=idx)
                cell.value = valor
                cell.font = Font(name='Calibri', size=10, bold=True)
                cell.border = self.border_thin
                if idx % 2 == 0:  # Colunas de valor
                    cell.alignment = self.align_right
                else:
                    cell.alignment = self.align_left
            self.linha_atual += 1
        
        # Ajustar larguras
        self.ajustar_largura_colunas([15, 35, 12, 15, 15, 18])


class RelatorioFluxoCaixaExcel(RelatorioExcelBase):
    """Relatório de fluxo de caixa em Excel"""
    
    def __init__(self, parcelas, dias, usuario):
        super().__init__()
        self.ws.title = "Fluxo de Caixa"
        
        # Cabeçalho
        titulo = "RELATÓRIO - FLUXO DE CAIXA PROJETADO"
        subtitulo = f"Próximos {dias} dias"
        self.adicionar_cabecalho(titulo, subtitulo, usuario)
        
        # Agrupar por semana com datas
        from collections import defaultdict
        from datetime import timedelta
        por_semana = {}
        
        hoje = date.today()
        for parcela in parcelas:
            dias_vencer = (parcela.data_vencimento - hoje).days
            num_semana = (dias_vencer // 7) + 1
            
            # Calcular data inicial e final da semana
            data_inicio = hoje + timedelta(days=(num_semana - 1) * 7)
            data_fim = data_inicio + timedelta(days=6)
            
            periodo = f"{self.formatar_data(data_inicio)} a {self.formatar_data(data_fim)}"
            
            if periodo not in por_semana:
                por_semana[periodo] = {'parcelas': [], 'ordem': num_semana}
            por_semana[periodo]['parcelas'].append(parcela)
        
        # Cabeçalho da tabela
        colunas = ['Período', 'Quantidade', 'Valor Total']
        self.adicionar_linha_cabecalho(colunas)
        
        # Dados ordenados por semana
        total_geral = Decimal('0')
        total_parcelas = 0
        
        for periodo in sorted(por_semana.keys(), key=lambda x: por_semana[x]['ordem']):
            parcelas_semana = por_semana[periodo]['parcelas']
            total_semana = sum(p.valor for p in parcelas_semana)
            
            valores = [
                periodo,
                len(parcelas_semana),
                self.formatar_moeda(total_semana)
            ]
            
            self.adicionar_linha_dados(valores)
            total_geral += total_semana
            total_parcelas += len(parcelas_semana)
        
        # Totais
        self.linha_atual += 1
        self.adicionar_linha_total('TOTAL DE PARCELAS:', str(total_parcelas), 1, 2)
        self.adicionar_linha_total('TOTAL PROJETADO:', self.formatar_moeda(total_geral), 1, 3)
        
        # Ajustar larguras (período mais largo para datas)
        self.ajustar_largura_colunas([35, 15, 20])


class RelatorioPorClienteExcel(RelatorioExcelBase):
    """Relatório por cliente em Excel"""
    
    def __init__(self, cliente, estatisticas, parcelas_pendentes, parcelas_vencidas, parcelas_pagas, usuario):
        super().__init__()
        self.ws.title = f"Cliente {cliente.id}"
        
        # Cabeçalho
        titulo = f"RELATÓRIO - {cliente.nome}"
        subtitulo = f"CPF/CNPJ: {cliente.cpf_cnpj}"
        self.adicionar_cabecalho(titulo, subtitulo, usuario)
        
        # Estatísticas
        self.ws.merge_cells(f'A{self.linha_atual}:F{self.linha_atual}')
        cell = self.ws[f'A{self.linha_atual}']
        cell.value = "RESUMO ESTATÍSTICO"
        cell.font = Font(name='Calibri', size=12, bold=True, color='0369A1')
        cell.alignment = self.align_center
        self.linha_atual += 1
        
        stats = [
            ['Total de Parcelas:', estatisticas['total_parcelas']],
            ['Pendentes:', estatisticas['pendentes']],
            ['Vencidas:', estatisticas['vencidas']],
            ['Pagas:', estatisticas['pagas']],
            ['Total Pendente:', self.formatar_moeda(estatisticas['total_pendente'])],
            ['Total Vencido:', self.formatar_moeda(estatisticas['total_vencido'])],
            ['Total Pago:', self.formatar_moeda(estatisticas['total_pago'])],
        ]
        
        for stat in stats:
            cell_label = self.ws.cell(row=self.linha_atual, column=1)
            cell_label.value = stat[0]
            cell_label.font = Font(name='Calibri', size=10, bold=True)
            
            cell_valor = self.ws.cell(row=self.linha_atual, column=2)
            cell_valor.value = stat[1]
            cell_valor.font = Font(name='Calibri', size=10)
            
            self.linha_atual += 1
        
        self.linha_atual += 1
        
        # Parcelas Pendentes
        if parcelas_pendentes:
            self.ws.merge_cells(f'A{self.linha_atual}:F{self.linha_atual}')
            cell = self.ws[f'A{self.linha_atual}']
            cell.value = "PARCELAS PENDENTES"
            cell.font = Font(name='Calibri', size=11, bold=True, color='F59E0B')
            cell.alignment = self.align_center
            self.linha_atual += 1
            
            colunas = ['NF', 'Parcela', 'Vencimento', 'Dias', 'Valor', 'Status']
            self.adicionar_linha_cabecalho(colunas)
            
            for parcela in parcelas_pendentes:
                dias_vencer = (parcela.data_vencimento - date.today()).days
                valores = [
                    parcela.nota_fiscal.numero_nota if parcela.nota_fiscal else '-',
                    parcela.numero_parcela,
                    self.formatar_data(parcela.data_vencimento),
                    dias_vencer,
                    self.formatar_moeda(parcela.valor),
                    'A Vencer' if dias_vencer > 0 else 'Vencida'
                ]
                destacar = 'vencido' if dias_vencer <= 0 else ('alerta' if dias_vencer <= 7 else None)
                self.adicionar_linha_dados(valores, destacar)
            
            self.linha_atual += 1
        
        # Parcelas Pagas Recentemente
        if parcelas_pagas:
            self.ws.merge_cells(f'A{self.linha_atual}:F{self.linha_atual}')
            cell = self.ws[f'A{self.linha_atual}']
            cell.value = "ÚLTIMAS PARCELAS PAGAS"
            cell.font = Font(name='Calibri', size=11, bold=True, color='10B981')
            cell.alignment = self.align_center
            self.linha_atual += 1
            
            colunas = ['NF', 'Parcela', 'Vencimento', 'Pagamento', 'Valor', 'Atraso']
            self.adicionar_linha_cabecalho(colunas)
            
            for parcela in parcelas_pagas[:10]:  # Últimas 10
                atraso = (parcela.data_pagamento - parcela.data_vencimento).days if parcela.data_pagamento else 0
                valores = [
                    parcela.nota_fiscal.numero_nota if parcela.nota_fiscal else '-',
                    parcela.numero_parcela,
                    self.formatar_data(parcela.data_vencimento),
                    self.formatar_data(parcela.data_pagamento),
                    self.formatar_moeda(parcela.valor),
                    f"{atraso} dias" if atraso > 0 else "Em dia"
                ]
                self.adicionar_linha_dados(valores)
        
        # Ajustar larguras
        self.ajustar_largura_colunas([15, 12, 15, 15, 18, 15])
